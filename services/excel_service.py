import openpyxl
from openpyxl import load_workbook
import os
import re
import logging
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional

# ============================================
# LOGGING CONFIGURATION
# ============================================
logger = logging.getLogger(__name__)

# ============================================
# CONSTANTS FOR HALF-DAY / HOURLY DETECTION
# ============================================

# Patterns for half-day leave detection
HALF_DAY_PATTERNS = [
    r'半',           # 半 = half
    r'0\.5',         # 0.5 explicit
    r'午前',         # 午前 = morning (AM half-day)
    r'午後',         # 午後 = afternoon (PM half-day)
    r'半休',         # 半休 = half-day leave
    r'AM',           # AM half-day
    r'PM',           # PM half-day
]

# Patterns for hourly leave detection (2 hours = 0.25 days)
HOURLY_PATTERNS = [
    r'2h',           # 2 hours
    r'2時間',        # 2時間 = 2 hours
    r'時間休',       # 時間休 = hourly leave
    r'(\d+)h',       # Generic hours pattern
    r'(\d+)時間',    # Generic hours in Japanese
]

# Patterns to remove when parsing dates
DATE_NOISE_PATTERNS = [
    r'終業',         # end of work
    r'至',           # until
    r'から',         # from
    r'まで',         # until
    r'休',           # leave
    r'取得',         # acquired
]


# ============================================
# HELPER FUNCTIONS FOR LEAVE TYPE DETECTION
# ============================================

def detect_leave_type(cell_value: Any) -> Tuple[str, float]:
    """
    Detects the type of leave from a cell value.

    Args:
        cell_value: The raw cell value (can be datetime, string, or number)

    Returns:
        Tuple of (leave_type, days_used)
        - leave_type: 'full', 'half_am', 'half_pm', 'hourly', or 'unknown'
        - days_used: float value (1.0, 0.5, 0.25, etc.)
    """
    if cell_value is None:
        return ('unknown', 0.0)

    cell_str = str(cell_value).strip()

    # Check for hourly leave first (more specific)
    for pattern in HOURLY_PATTERNS:
        match = re.search(pattern, cell_str, re.IGNORECASE)
        if match:
            # Try to extract hours if pattern has group
            if match.groups():
                try:
                    hours = int(match.group(1))
                    # Assuming 8-hour workday
                    days = hours / 8.0
                    return ('hourly', round(days, 2))
                except (ValueError, IndexError):
                    pass
            # Default 2 hours = 0.25 days
            return ('hourly', 0.25)

    # Check for half-day patterns
    for pattern in HALF_DAY_PATTERNS:
        if re.search(pattern, cell_str, re.IGNORECASE):
            # Determine AM or PM
            if re.search(r'午前|AM', cell_str, re.IGNORECASE):
                return ('half_am', 0.5)
            elif re.search(r'午後|PM', cell_str, re.IGNORECASE):
                return ('half_pm', 0.5)
            else:
                return ('half_am', 0.5)  # Default to AM if not specified

    # Check if it's just a date (full day)
    if isinstance(cell_value, datetime):
        return ('full', 1.0)

    # If it's a string that looks like a date, it's a full day
    if cell_str and not any(pattern in cell_str.lower() for pattern in ['半', '0.5', 'h', '時間']):
        # Check if it contains date-like patterns
        if re.search(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', cell_str):
            return ('full', 1.0)

    return ('full', 1.0)  # Default to full day


def clean_date_string(date_str: str) -> str:
    """
    Removes noise patterns from a date string to facilitate parsing.

    Args:
        date_str: Raw date string possibly containing Japanese text

    Returns:
        Cleaned date string
    """
    cleaned = date_str
    for pattern in DATE_NOISE_PATTERNS:
        cleaned = re.sub(pattern, '', cleaned)

    # Also remove common separators that aren't part of dates
    cleaned = cleaned.strip()

    return cleaned


def parse_date_from_cell(cell_value: Any) -> Optional[datetime]:
    """
    Attempts to parse a date from various cell formats.
    Handles comments and mixed content gracefully.

    Args:
        cell_value: The cell value (datetime, string, number)

    Returns:
        datetime object if successfully parsed, None otherwise
    """
    if cell_value is None:
        return None

    # Already a datetime
    if isinstance(cell_value, datetime):
        return cell_value

    # Excel serial date (number)
    if isinstance(cell_value, (int, float)):
        try:
            from datetime import timedelta
            # Excel epoch is 1899-12-30
            excel_epoch = datetime(1899, 12, 30)
            # Validate reasonable range (1900-2100)
            if 1 < cell_value < 73050:  # ~2100
                return excel_epoch + timedelta(days=int(cell_value))
        except Exception:
            pass
        return None

    # String parsing
    if isinstance(cell_value, str):
        date_str = clean_date_string(cell_value)

        # Try various date formats
        date_formats = [
            '%Y/%m/%d',
            '%Y-%m-%d',
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%d %H:%M:%S',
            '%Y年%m月%d日',
            '%m/%d/%Y',
            '%d/%m/%Y',
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue

        # Try to extract date using regex
        date_match = re.search(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', date_str)
        if date_match:
            try:
                year = int(date_match.group(1))
                month = int(date_match.group(2))
                day = int(date_match.group(3))
                return datetime(year, month, day)
            except ValueError:
                pass

    return None


def is_valid_date_for_yukyu(parsed_date: Optional[datetime]) -> bool:
    """
    Validates if a parsed date is reasonable for yukyu (paid leave) records.

    Args:
        parsed_date: The datetime to validate

    Returns:
        True if the date is valid, False otherwise
    """
    if parsed_date is None:
        return False

    # Reasonable year range: 2000-2100
    if parsed_date.year < 2000 or parsed_date.year > 2100:
        return False

    return True

def parse_excel_file(file_path):
    """
    Parses an Excel file and returns a list of employee dictionaries.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Match JS logic: Use the first sheet
    sheet = wb.worksheets[0]
    
    data = []
    
    # Iterate to find header row (looking for '氏名' or '名前')
    header_row_idx = 1
    headers = []
    
    # Scan first 10 rows for headers
    found_headers = False
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_str = [str(c) if c is not None else "" for c in row]
        if any("氏名" in c or "名前" in c for c in row_str):
            headers = row
            header_row_idx = i
            found_headers = True
            break
            
    if not found_headers:
        # Fallback to row 1 if detection fails
        headers = [cell.value for cell in sheet[1]]
        header_row_idx = 1

    # Identify column indices
    cols = {
        'empNum': _find_obs_col(headers, ['社員№', '従業員番号', '社員番号', '番号', 'id', 'no', '№']),
        'name': _find_obs_col(headers, ['氏名', '名前', 'name']),
        'haken': _find_obs_col(headers, ['派遣先', '所属', '部署', '現場']),
        'granted': _find_obs_col(headers, ['付与数', '付与日数', '付与', '総日数', '有給残日数']),
        'used': _find_obs_col(headers, ['消化日数', '使用日数', '使用', '消化']),
        'balance': _find_obs_col(headers, ['期末残高', 'balance', 'remaining', '残高']),
        'expired': _find_obs_col(headers, ['時効数', 'expired', '期限切れ', '消滅']),
        'year': _find_obs_col(headers, ['年度', '年', 'year', '対象年度']),
        'grant_date': _find_obs_col(headers, ['有給発生', '発生日', '付与日'])
    }
    
    # Default year if column not found
    current_year = datetime.now().year

    # ACCUMULATE BY EMPLOYEE (like v2.0) - Multiple rows for same employee are summed
    employee_summary = {}

    # Start reading from the row AFTER headers
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(c is None for c in row):
            continue

        emp_num = str(row[cols['empNum']]) if cols['empNum'] != -1 and row[cols['empNum']] is not None else "Unknown"
        name = str(row[cols['name']]) if cols['name'] != -1 and row[cols['name']] is not None else "Unknown"

        # Skip if name is missing (critical for dashboard)
        if name == "Unknown":
            continue

        # Skip duplicate header rows (where 付与 column contains text like "付与数")
        if cols['granted'] != -1 and row[cols['granted']] is not None:
            granted_value = str(row[cols['granted']])
            if any(header_text in granted_value for header_text in ['付与', '日数', 'granted']):
                continue

        haken = str(row[cols['haken']]) if cols['haken'] != -1 and row[cols['haken']] is not None else "Unknown"

        # Parse numeric values
        try:
            granted = float(row[cols['granted']]) if cols['granted'] != -1 and row[cols['granted']] is not None else 0.0
            used = float(row[cols['used']]) if cols['used'] != -1 and row[cols['used']] is not None else 0.0
        except (ValueError, TypeError):
            granted = 0.0
            used = 0.0

        # Read balance from Excel (CRITICAL: Do NOT calculate it)
        try:
            if cols['balance'] != -1 and row[cols['balance']] is not None:
                balance = float(row[cols['balance']])
            else:
                # Fallback to calculation ONLY if Excel doesn't have balance column
                balance = granted - used
        except (ValueError, TypeError):
            balance = granted - used

        # Read expired days from Excel
        try:
            expired = float(row[cols['expired']]) if cols['expired'] != -1 and row[cols['expired']] is not None else 0.0
        except (ValueError, TypeError):
            expired = 0.0

        # Try to extract year from grant_date first (有給発生 column)
        year = current_year
        if cols['grant_date'] != -1 and row[cols['grant_date']] is not None:
            grant_date_value = row[cols['grant_date']]
            try:
                # If it's already a datetime object
                if hasattr(grant_date_value, 'year'):
                    year = grant_date_value.year
                # If it's a string, try to parse it
                elif isinstance(grant_date_value, str):
                    # Try to extract year from date string
                    parsed_date = datetime.strptime(grant_date_value.split()[0], '%Y-%m-%d')
                    year = parsed_date.year
            except (ValueError, AttributeError, IndexError):
                pass

        # Fallback to year column if grant_date didn't work
        if year == current_year and cols['year'] != -1 and row[cols['year']] is not None:
            try:
                year = int(row[cols['year']])
            except ValueError:
                pass

        # Create unique key: empNum + year
        emp_key = f"{emp_num}_{year}"

        # NO ACCUMULATION - Each row with same empNum+year REPLACES the previous
        # This prevents double counting when employee has multiple grant periods
        # that might be incorrectly assigned to the same year
        employee_summary[emp_key] = {
            'id': emp_key,
            'employeeNum': emp_num,
            'name': name,
            'haken': haken,
            'granted': granted,
            'used': used,
            'balance': balance,
            'expired': expired,
            'year': year
        }

    # Convert summary dict to list and calculate usage rates
    for emp in employee_summary.values():
        emp['usageRate'] = round((emp['used'] / emp['granted']) * 100) if emp['granted'] > 0 else 0
        data.append(emp)

    return data

def _find_obs_col(headers, keywords):
    """Finds the index of a column matching one of the keywords.
    Searches keywords in order, finding the first match across all headers."""
    for k in keywords:
        for i, h in enumerate(headers):
            if h is None: continue
            h_str = str(h).lower()
            k_lower = str(k).lower()
            if k_lower in h_str:
                return i
    return -1


def parse_genzai_sheet(file_path):
    """
    Parses DBGenzaiX sheet from the employee registry Excel file.
    Returns a list of dispatch employee dictionaries.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Look for DBGenzaiX sheet
    if 'DBGenzaiX' not in wb.sheetnames:
        raise ValueError("DBGenzaiX sheet not found in workbook")

    sheet = wb['DBGenzaiX']
    data = []

    # Expected headers (row 1)
    # 0:現在, 1:社員№, 2:派遣先ID, 3:派遣先, 4:配属先, 5:配属ライン, 6:仕事内容,
    # 7:氏名, 8:カナ, 9:性別, 10:国籍, 11:生年月日, 12:年齢, 13:時給, 14:時給改定

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if all(c is None for c in row):
            continue

        # Extract core fields
        status = str(row[0]) if row[0] is not None else "Unknown"
        emp_num = str(row[1]) if row[1] is not None else "Unknown"
        name = str(row[7]) if row[7] is not None else "Unknown"

        # Skip if no employee number or name, or if they are "0"
        if emp_num == "Unknown" or name == "Unknown" or emp_num == "0" or name == "0":
            continue

        # Parse birth date
        birth_date = None
        if row[11] is not None:
            if isinstance(row[11], datetime):
                birth_date = row[11].strftime('%Y-%m-%d')
            else:
                birth_date = str(row[11])

        # Generate ID
        emp_id = f"genzai_{emp_num}"

        employee = {
            'id': emp_id,
            'status': status,
            'employee_num': emp_num,
            'dispatch_id': str(row[2]) if row[2] is not None else "",
            'dispatch_name': str(row[3]) if row[3] is not None else "",
            'department': str(row[4]) if row[4] is not None else "",
            'line': str(row[5]) if row[5] is not None else "",
            'job_content': str(row[6]) if row[6] is not None else "",
            'name': name,
            'kana': str(row[8]) if row[8] is not None else "",
            'gender': str(row[9]) if row[9] is not None else "",
            'nationality': str(row[10]) if row[10] is not None else "",
            'birth_date': birth_date,
            'age': int(row[12]) if row[12] is not None else None,
            'hourly_wage': int(row[13]) if row[13] is not None else 0,
            'wage_revision': str(row[14]) if row[14] is not None else ""
        }
        data.append(employee)

    return data


def parse_ukeoi_sheet(file_path):
    """
    Parses DBUkeoiX sheet from the employee registry Excel file.
    Returns a list of contract employee dictionaries.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Look for DBUkeoiX sheet
    if 'DBUkeoiX' not in wb.sheetnames:
        raise ValueError("DBUkeoiX sheet not found in workbook")

    sheet = wb['DBUkeoiX']
    data = []

    # Expected headers (row 1)
    # 0:現在, 1:社員№, 2:請負業務, 3:氏名, 4:カナ, 5:性別, 6:国籍,
    # 7:生年月日, 8:年齢, 9:時給, 10:時給改定

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if all(c is None for c in row):
            continue

        # Extract core fields
        status = str(row[0]) if row[0] is not None else "Unknown"
        emp_num = str(row[1]) if row[1] is not None else "Unknown"
        name = str(row[3]) if row[3] is not None else "Unknown"

        # Skip if no employee number or name, or if they are "0"
        if emp_num == "Unknown" or name == "Unknown" or emp_num == "0" or name == "0":
            continue

        # Parse birth date
        birth_date = None
        if row[7] is not None:
            if isinstance(row[7], datetime):
                birth_date = row[7].strftime('%Y-%m-%d')
            else:
                birth_date = str(row[7])

        # Generate ID
        emp_id = f"ukeoi_{emp_num}"

        employee = {
            'id': emp_id,
            'status': status,
            'employee_num': emp_num,
            'contract_business': str(row[2]) if row[2] is not None else "",
            'name': name,
            'kana': str(row[4]) if row[4] is not None else "",
            'gender': str(row[5]) if row[5] is not None else "",
            'nationality': str(row[6]) if row[6] is not None else "",
            'birth_date': birth_date,
            'age': int(row[8]) if row[8] is not None else None,
            'hourly_wage': int(row[9]) if row[9] is not None else 0,
            'wage_revision': str(row[10]) if row[10] is not None else ""
        }
        data.append(employee)

    return data


def parse_staff_sheet(file_path):
    """
    Parses the DBStaffX sheet from the employee registry Excel file.
    Includes hire_date (入社日) and leave_date (退社日) for year filtering.

    Expected columns:
    0:№(status), 1:社員№, 2:事務所, 3:氏名, 4:カナ, 5:性別, 6:国籍,
    7:生年月日, 8:年齢, 9:ビザ期限, 10:ビザ種類, 11:配偶者,
    12:〒, 13:住所, 14:建物名, 15:入社日, 16:退社日

    Returns:
        List of staff employee dicts
    """
    wb = load_workbook(file_path, data_only=True)

    if 'DBStaffX' not in wb.sheetnames:
        raise ValueError("DBStaffX sheet not found in workbook")

    sheet = wb['DBStaffX']
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if all(c is None for c in row):
            continue

        # Extract core fields
        status = str(row[0]) if row[0] is not None else "Unknown"
        emp_num = str(row[1]) if row[1] is not None else "Unknown"
        name = str(row[3]) if row[3] is not None else "Unknown"

        # Skip if no employee number or name
        if emp_num == "Unknown" or name == "Unknown" or emp_num == "0" or name == "0":
            continue

        # Parse dates
        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.strftime('%Y-%m-%d')
            if isinstance(val, (int, float)) and val > 0:
                # Excel serial date
                try:
                    from datetime import timedelta
                    excel_epoch = datetime(1899, 12, 30)
                    return (excel_epoch + timedelta(days=int(val))).strftime('%Y-%m-%d')
                except:
                    return None
            return str(val) if val else None

        birth_date = parse_date(row[7])
        visa_expiry = parse_date(row[9])
        hire_date = parse_date(row[15])
        leave_date = parse_date(row[16])

        emp_id = f"staff_{emp_num}"

        employee = {
            'id': emp_id,
            'status': status,
            'employee_num': emp_num,
            'office': str(row[2]) if row[2] is not None else "",
            'name': name,
            'kana': str(row[4]) if row[4] is not None else "",
            'gender': str(row[5]) if row[5] is not None else "",
            'nationality': str(row[6]) if row[6] is not None else "",
            'birth_date': birth_date,
            'age': int(row[8]) if row[8] is not None else None,
            'visa_expiry': visa_expiry,
            'visa_type': str(row[10]) if row[10] is not None else "",
            'spouse': str(row[11]) if row[11] is not None else "",
            'postal_code': str(row[12]) if row[12] is not None else "",
            'address': str(row[13]) if row[13] is not None else "",
            'building': str(row[14]) if row[14] is not None else "",
            'hire_date': hire_date,
            'leave_date': leave_date
        }
        data.append(employee)

    return data


def parse_yukyu_usage_details(file_path):
    """
    Parses individual yukyu usage dates from columns R-BE (18-57).
    This extracts the INDIVIDUAL DATES when each employee used yukyu (like v2.0).
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        List of dicts with individual usage details:
        [
            {
                'employee_num': '123',
                'name': 'John Doe',
                'use_date': '2025-03-15',
                'year': 2025,
                'month': 3,
                'days_used': 1.0
            },
            ...
        ]
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    usage_details = []
    
    # Header row is at row 5
    header_row_idx = 5
    
    # Find employee info columns
    emp_num_col = 3   # Column C: 社員番号
    name_col = 5      # Column E: 氏名
    
    # Date columns are R to BE (columns 18-57 = 40 columns)
    date_start_col = 18  # Column R
    date_end_col = 57    # Column BE
    
    # Process each data row (starting after header)
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        # Get employee info
        emp_num = sheet.cell(row=row_idx, column=emp_num_col).value
        name = sheet.cell(row=row_idx, column=name_col).value
        
        # Skip if no employee number or name
        if not emp_num or not name:
            continue
        
        emp_num = str(emp_num)
        name = str(name)
        
        # Extract individual usage dates from columns R-BE
        for col_idx in range(date_start_col, date_end_col + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            
            # Skip empty cells
            if cell_value is None:
                continue
            
            # Parse date
            use_date = None
            
            if isinstance(cell_value, datetime):
                # It's already a datetime object
                use_date = cell_value.strftime('%Y-%m-%d')
                year = cell_value.year
                month = cell_value.month
            elif isinstance(cell_value, str):
                # Try to parse string dates
                # Format examples: "2025/7/1終業", "2025-3-1至1", etc.
                try:
                    # Remove Japanese text markers
                    date_str = cell_value.replace('終業', '').replace('至', '').strip()
                    
                    # Try various date formats
                    for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S']:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            use_date = parsed_date.strftime('%Y-%m-%d')
                            year = parsed_date.year
                            month = parsed_date.month
                            break
                        except ValueError:
                            continue
                except:
                    continue
            
            # If we successfully parsed a date, add it
            if use_date:
                usage_details.append({
                    'employee_num': emp_num,
                    'name': name,
                    'use_date': use_date,
                    'year': year,
                    'month': month,
                    'days_used': 1.0  # Default to 1 day per date
                })
    
    wb.close()
    return usage_details


def parse_yukyu_usage_details_enhanced(file_path: str) -> Dict[str, Any]:
    """
    Enhanced version of parse_yukyu_usage_details using Pandas for robustness.
    - Automatic half-day detection (半, 0.5, 午前, 午後)
    - Hourly leave detection (2h, 2時間)
    - Comment handling
    - Robust date parsing
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dict with keys: data, warnings, errors, summary
    """
    logger.info(f"Starting enhanced parsing (Pandas) of: {file_path}")
    
    result = {
        'data': [],
        'warnings': [],
        'errors': [],
        'summary': {
            'total_rows_processed': 0,
            'total_dates_parsed': 0,
            'half_day_count': 0,
            'hourly_count': 0,
            'full_day_count': 0
        }
    }

    if not os.path.exists(file_path):
        result['errors'].append(f"File not found: {file_path}")
        return result

    try:
        # Determine sheet name
        xl = pd.ExcelFile(file_path)
        sheet_name = 0
        if '作業者データ　有給' in xl.sheet_names:
            sheet_name = '作業者データ　有給'
            logger.info("Targeting sheet: 作業者データ　有給")
        else:
            logger.info(f"Targeting first sheet: {xl.sheet_names[0]}")

        # Read Excel with Pandas
        # Header is usually row 5 (0-indexed = 4)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)
        
        # Identify Name Column (usually index 4 / '氏名')
        name_col = None
        for col in df.columns:
            if '氏名' in str(col) or '名前' in str(col):
                name_col = col
                break
        
        if name_col is None:
            # Fallback to index 4 if column name not found
            if len(df.columns) > 4:
                name_col = df.columns[4]
            else:
                result['errors'].append("Could not identify Name column")
                return result
                
        # Employee Num Column (usually index 2 or 3)
        emp_num_col = None
        for col in df.columns:
            if '社員' in str(col) and ('番号' in str(col) or 'No' in str(col) or '№' in str(col)):
                emp_num_col = col
                break
        
        if emp_num_col is None:
             if len(df.columns) > 2:
                 emp_num_col = df.columns[2]

        result['summary']['total_rows_processed'] = len(df)

        # Date columns range (R=17 to BE=56 approx)
        # We'll take slice from 17 to 57
        date_columns = df.columns[17:57]

        usage_details = []

        for idx, row in df.iterrows():
            name = row[name_col]
            emp_num = row[emp_num_col] if emp_num_col else "Unknown"

            if pd.isna(name) or str(name).strip() == "":
                continue

            name = str(name).strip()
            emp_num = str(emp_num).strip()

            # Iterate through date columns
            for col in date_columns:
                cell_value = row[col]
                
                if pd.isna(cell_value) or cell_value == "":
                    continue

                # Parse Cell Value
                days_used = 1.0
                leave_type = 'full'
                use_date_str = None
                
                val_str = str(cell_value).strip()
                
                # Detect Half Day / Hourly
                # Reuse logic from helper or inline simplified
                if any(x in val_str for x in ['半', '0.5', 'AM', 'PM', '午前', '午後']):
                    days_used = 0.5
                    leave_type = 'half'
                    result['summary']['half_day_count'] += 1
                elif '2h' in val_str or '2時間' in val_str:
                    days_used = 0.25
                    leave_type = 'hourly'
                    result['summary']['hourly_count'] += 1
                else:
                    result['summary']['full_day_count'] += 1

                # Parse Date
                # If it's a datetime object (Pandas often converts automatically)
                if isinstance(cell_value, datetime):
                    use_date_str = cell_value.strftime('%Y-%m-%d')
                    year = cell_value.year
                    month = cell_value.month
                else:
                    # String parsing
                    # Clean the string first (remove '終業', '半', etc)
                    clean_str = re.sub(r'[^\d\/\-\.]', '', val_str)
                    
                    try:
                        # Try common formats
                        dt = pd.to_datetime(clean_str, errors='coerce')
                        if not pd.isna(dt):
                           use_date_str = dt.strftime('%Y-%m-%d')
                           year = dt.year
                           month = dt.month
                    except:
                        pass
                
                if use_date_str:
                    usage_details.append({
                        'employee_num': emp_num,
                        'name': name,
                        'use_date': use_date_str,
                        'year': year,
                        'month': month,
                        'days_used': days_used
                    })
                    result['summary']['total_dates_parsed'] += 1
                else:
                    # Could not parse date
                    pass # unexpected format

        result['data'] = usage_details
        logger.info(f"Pandas parsing complete. Found {len(usage_details)} usage records.")

    except Exception as e:
        logger.error(f"Pandas parsing failed: {e}", exc_info=True)
        result['errors'].append(f"Pandas Parsing Error: {str(e)}")

    return result



def get_import_report(enhanced_result: Dict[str, Any]) -> str:
    """
    Generates a human-readable import report from enhanced parsing results.

    Args:
        enhanced_result: Result from parse_yukyu_usage_details_enhanced()

    Returns:
        Formatted string report
    """
    summary = enhanced_result.get('summary', {})
    warnings = enhanced_result.get('warnings', [])
    errors = enhanced_result.get('errors', [])

    report_lines = [
        "=" * 60,
        "YUKYU IMPORT REPORT",
        "=" * 60,
        "",
        "STATISTICS:",
        f"  Total rows processed: {summary.get('total_rows_processed', 0)}",
        f"  Total dates parsed:   {summary.get('total_dates_parsed', 0)}",
        f"    - Full days:        {summary.get('full_day_count', 0)}",
        f"    - Half days:        {summary.get('half_day_count', 0)}",
        f"    - Hourly:           {summary.get('hourly_count', 0)}",
        f"  Dates ignored:        {summary.get('dates_ignored', 0)}",
        f"  Invalid dates:        {summary.get('invalid_dates', 0)}",
        f"  Cells with comments:  {summary.get('cells_with_comments', 0)}",
        "",
    ]

    if errors:
        report_lines.extend([
            "ERRORS:",
            "-" * 40,
        ])
        for error in errors[:10]:  # Limit to first 10
            if isinstance(error, dict):
                report_lines.append(f"  - Row {error.get('row')}: {error.get('message')}")
            else:
                report_lines.append(f"  - {error}")
        if len(errors) > 10:
            report_lines.append(f"  ... and {len(errors) - 10} more errors")
        report_lines.append("")

    if warnings:
        report_lines.extend([
            "WARNINGS (first 20):",
            "-" * 40,
        ])
        for warning in warnings[:20]:  # Limit to first 20
            if isinstance(warning, dict):
                report_lines.append(
                    f"  [{warning.get('type', 'unknown')}] Row {warning.get('row')}, "
                    f"Employee {warning.get('employee')}: {warning.get('message')}"
                )
            else:
                report_lines.append(f"  - {warning}")
        if len(warnings) > 20:
            report_lines.append(f"  ... and {len(warnings) - 20} more warnings")
        report_lines.append("")

    report_lines.extend([
        "=" * 60,
        "END OF REPORT",
        "=" * 60,
    ])

    return "\n".join(report_lines)
