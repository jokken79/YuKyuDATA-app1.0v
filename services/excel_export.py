"""
Módulo de exportación a Excel para YuKyuDATA-app

Funcionalidades:
- Exportar solicitudes aprobadas a Excel
- Generar reportes mensuales en Excel
- Actualizar archivo maestro con datos de la BD
- Sincronización bidireccional
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from typing import Optional, List, Dict
import os
import sqlite3
import shutil

# Directorio de exportaciones
EXPORT_DIR = os.path.join(os.path.dirname(__file__), 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)

DB_NAME = "yukyu.db"


from contextlib import contextmanager

@contextmanager
def get_db_connection():
    """
    Obtiene conexión a la base de datos como context manager.

    Uso:
        with get_db_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT ...")
    """
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


# Estilos reutilizables
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

LEAVE_TYPE_NAMES = {
    'full': '全日',
    'half_am': '午前半休',
    'half_pm': '午後半休',
    'hourly': '時間休'
}


def apply_header_style(cell):
    """Aplica estilo de header a una celda"""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER
    cell.alignment = CENTER_ALIGN


def apply_cell_style(cell, align='left'):
    """Aplica estilo de celda de datos"""
    cell.border = BORDER
    if align == 'center':
        cell.alignment = CENTER_ALIGN
    elif align == 'right':
        cell.alignment = RIGHT_ALIGN


def create_approved_requests_excel(year: int, month: int = None) -> str:
    """
    Crea Excel con solicitudes aprobadas.

    Args:
        year: Año a exportar
        month: Mes específico (opcional)

    Returns:
        Ruta del archivo creado
    """
    with get_db_connection() as conn:
        c = conn.cursor()

        query = '''
            SELECT
                lr.id,
                lr.employee_num,
                lr.employee_name,
                lr.start_date,
                lr.end_date,
                lr.days_requested,
                lr.hours_requested,
                lr.leave_type,
                lr.reason,
                lr.approved_at,
                lr.approved_by,
                lr.cost_estimate
            FROM leave_requests lr
            WHERE lr.status = 'APPROVED'
            AND strftime('%Y', lr.approved_at) = ?
        '''
        params = [str(year)]

        if month:
            query += " AND strftime('%m', lr.approved_at) = ?"
            params.append(f"{month:02d}")

        query += " ORDER BY lr.approved_at DESC"

        requests = c.execute(query, params).fetchall()

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "承認済み申請"

    # Headers
    headers = [
        'ID', '社員番号', '氏名', '開始日', '終了日',
        '日数', '時間', '種類', '理由', '承認日', '承認者', '費用見積(¥)'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # Datos
    total_days = 0
    total_hours = 0
    total_cost = 0

    for row_idx, req in enumerate(requests, 2):
        days = float(req['days_requested']) if req['days_requested'] else 0
        hours = float(req['hours_requested']) if req['hours_requested'] else 0
        cost = float(req['cost_estimate']) if req['cost_estimate'] else 0

        total_days += days
        total_hours += hours
        total_cost += cost

        data = [
            req['id'],
            req['employee_num'],
            req['employee_name'],
            req['start_date'],
            req['end_date'],
            days,
            hours,
            LEAVE_TYPE_NAMES.get(req['leave_type'], req['leave_type']),
            req['reason'] or '',
            req['approved_at'][:10] if req['approved_at'] else '',
            req['approved_by'] or '',
            cost
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            align = 'right' if col in [1, 6, 7, 12] else 'center' if col in [4, 5, 10] else 'left'
            apply_cell_style(cell, align)

    # Fila de totales
    total_row = len(requests) + 2
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_days).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_hours).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=total_cost).font = Font(bold=True)

    # Ajustar anchos de columna
    column_widths = [8, 12, 15, 12, 12, 8, 8, 12, 30, 12, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Guardar
    filename = f"approved_requests_{year}"
    if month:
        filename += f"_{month:02d}"
    filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath


def create_monthly_report_excel(year: int, month: int) -> str:
    """
    Crea reporte mensual en Excel (período 21日〜20日).

    Args:
        year: Año
        month: Mes

    Returns:
        Ruta del archivo creado
    """
    conn = get_db_connection()
    c = conn.cursor()

    # Calcular período (21日〜20日)
    if month == 1:
        start_year = year - 1
        start_month = 12
    else:
        start_year = year
        start_month = month - 1

    start_date = f"{start_year}-{start_month:02d}-21"
    end_date = f"{year}-{month:02d}-20"

    # Obtener datos de uso detallado
    usage_data = c.execute('''
        SELECT
            employee_num,
            name,
            use_date,
            days_used
        FROM yukyu_usage_details
        WHERE use_date BETWEEN ? AND ?
        ORDER BY use_date, name
    ''', (start_date, end_date)).fetchall()

    # Obtener solicitudes aprobadas en el período
    approved = c.execute('''
        SELECT
            employee_num,
            employee_name,
            start_date,
            end_date,
            days_requested,
            hours_requested,
            leave_type,
            approved_at
        FROM leave_requests
        WHERE status = 'APPROVED'
        AND (
            (start_date BETWEEN ? AND ?)
            OR (end_date BETWEEN ? AND ?)
        )
        ORDER BY start_date
    ''', (start_date, end_date, start_date, end_date)).fetchall()

    # Resumen por empleado
    emp_summary = c.execute('''
        SELECT
            employee_num,
            name,
            SUM(days_used) as total_days,
            COUNT(*) as usage_count
        FROM yukyu_usage_details
        WHERE use_date BETWEEN ? AND ?
        GROUP BY employee_num
        ORDER BY total_days DESC
    ''', (start_date, end_date)).fetchall()

    conn.close()

    # Crear workbook
    wb = Workbook()

    # === Hoja 1: Resumen ===
    ws1 = wb.active
    ws1.title = "月次サマリー"

    # Título
    ws1['A1'] = f"{year}年{month}月度 有給休暇レポート"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')

    ws1['A2'] = f"対象期間: {start_date} 〜 {end_date}"
    ws1['A2'].font = Font(size=12, color="666666")

    # Estadísticas generales
    total_days_used = sum(float(u['days_used']) for u in usage_data)
    unique_employees = len(set(u['employee_num'] for u in usage_data))

    ws1['A4'] = "統計サマリー"
    ws1['A4'].font = Font(bold=True, size=12)

    stats = [
        ("取得者数", f"{unique_employees}人"),
        ("総取得日数", f"{total_days_used:.1f}日"),
        ("平均取得日数", f"{total_days_used/unique_employees:.1f}日" if unique_employees else "0日"),
        ("承認済み申請数", f"{len(approved)}件"),
    ]

    for i, (label, value) in enumerate(stats, 5):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=value)

    # Top 10 usuarios
    ws1['A10'] = "取得日数 Top 10"
    ws1['A10'].font = Font(bold=True, size=12)

    headers = ['順位', '社員番号', '氏名', '取得日数', '取得回数']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=11, column=col, value=header)
        apply_header_style(cell)

    for i, emp in enumerate(emp_summary[:10], 1):
        row = 11 + i
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=emp['employee_num'])
        ws1.cell(row=row, column=3, value=emp['name'])
        ws1.cell(row=row, column=4, value=float(emp['total_days']))
        ws1.cell(row=row, column=5, value=emp['usage_count'])

    # === Hoja 2: Detalle por fecha ===
    ws2 = wb.create_sheet("日別詳細")

    headers = ['使用日', '社員番号', '氏名', '日数']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    for row_idx, usage in enumerate(usage_data, 2):
        ws2.cell(row=row_idx, column=1, value=usage['use_date'])
        ws2.cell(row=row_idx, column=2, value=usage['employee_num'])
        ws2.cell(row=row_idx, column=3, value=usage['name'])
        ws2.cell(row=row_idx, column=4, value=float(usage['days_used']))

    # === Hoja 3: Solicitudes aprobadas ===
    ws3 = wb.create_sheet("承認済み申請")

    headers = ['社員番号', '氏名', '開始日', '終了日', '日数', '時間', '種類', '承認日']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    for row_idx, req in enumerate(approved, 2):
        ws3.cell(row=row_idx, column=1, value=req['employee_num'])
        ws3.cell(row=row_idx, column=2, value=req['employee_name'])
        ws3.cell(row=row_idx, column=3, value=req['start_date'])
        ws3.cell(row=row_idx, column=4, value=req['end_date'])
        ws3.cell(row=row_idx, column=5, value=float(req['days_requested']) if req['days_requested'] else 0)
        ws3.cell(row=row_idx, column=6, value=float(req['hours_requested']) if req['hours_requested'] else 0)
        ws3.cell(row=row_idx, column=7, value=LEAVE_TYPE_NAMES.get(req['leave_type'], req['leave_type']))
        ws3.cell(row=row_idx, column=8, value=req['approved_at'][:10] if req['approved_at'] else '')

    # Ajustar anchos
    for ws in [ws1, ws2, ws3]:
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Guardar
    filename = f"monthly_report_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath


def create_annual_ledger_excel(year: int) -> str:
    """
    Crea el libro de gestión anual de vacaciones (年次有給休暇管理簿).
    Requerido por ley japonesa.

    Args:
        year: Año fiscal

    Returns:
        Ruta del archivo creado
    """
    conn = get_db_connection()
    c = conn.cursor()

    # Obtener todos los empleados del año
    employees = c.execute('''
        SELECT
            e.employee_num,
            e.name,
            e.granted,
            e.used,
            e.balance,
            g.hire_date,
            g.dispatch_name as factory
        FROM employees e
        LEFT JOIN genzai g ON e.employee_num = g.employee_num
        WHERE e.year = ?
        ORDER BY e.name
    ''', (year,)).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "年次有給休暇管理簿"

    # Título
    ws['A1'] = f"{year}年度 年次有給休暇管理簿"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')

    ws['A2'] = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"

    # Headers
    headers = ['社員番号', '氏名', '入社日', '派遣先', '付与日数', '取得日数', '残日数', '取得率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)

    # Datos
    for row_idx, emp in enumerate(employees, 5):
        granted = float(emp['granted']) if emp['granted'] else 0
        used = float(emp['used']) if emp['used'] else 0
        balance = float(emp['balance']) if emp['balance'] else 0
        rate = round((used / granted * 100), 1) if granted > 0 else 0

        data = [
            emp['employee_num'],
            emp['name'],
            emp['hire_date'] or '-',
            emp['factory'] or '-',
            granted,
            used,
            balance,
            f"{rate}%"
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            apply_cell_style(cell)

    # Ajustar anchos
    column_widths = [12, 15, 12, 20, 10, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Guardar
    filename = f"annual_ledger_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath


def update_master_excel(master_path: str, year: int) -> Dict:
    """
    Actualiza el archivo Excel maestro con datos de la BD.
    Sincronización bidireccional: BD → Excel

    Args:
        master_path: Ruta al archivo Excel maestro
        year: Año a actualizar

    Returns:
        Dict con estadísticas de actualización
    """
    if not os.path.exists(master_path):
        return {"status": "error", "message": f"Archivo no encontrado: {master_path}"}

    conn = get_db_connection()
    c = conn.cursor()

    # Obtener datos actualizados de la BD
    employees = c.execute('''
        SELECT employee_num, name, granted, used, balance, usage_rate
        FROM employees
        WHERE year = ?
    ''', (year,)).fetchall()

    conn.close()

    # Crear backup antes de modificar
    backup_dir = os.path.join(os.path.dirname(master_path), 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    backup_filename = os.path.basename(master_path).replace(
        '.xlsm',
        f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
    )
    backup_path = os.path.join(backup_dir, backup_filename)

    try:
        shutil.copy2(master_path, backup_path)
    except Exception as e:
        return {"status": "error", "message": f"Error creando backup: {str(e)}"}

    try:
        # Cargar workbook preservando macros
        wb = load_workbook(master_path, keep_vba=True)
        ws = wb.active

        # Buscar fila de headers
        header_row = None
        col_map = {}

        for row in range(1, 15):
            for col in range(1, 30):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '氏名' in str(cell_value):
                    header_row = row
                    break
            if header_row:
                break

        if not header_row:
            return {"status": "error", "message": "No se encontró fila de headers"}

        # Mapear columnas
        for col in range(1, 60):
            val = ws.cell(row=header_row, column=col).value
            if val:
                val_str = str(val)
                if '従業員番号' in val_str or '社員番号' in val_str or 'No' in val_str:
                    col_map['employee_num'] = col
                elif '使用' in val_str and '日数' in val_str:
                    col_map['used'] = col
                elif '残' in val_str or 'バランス' in val_str or '残日数' in val_str:
                    col_map['balance'] = col
                elif '消化率' in val_str or '取得率' in val_str:
                    col_map['usage_rate'] = col

        if 'employee_num' not in col_map:
            return {"status": "error", "message": "No se encontró columna de número de empleado"}

        # Actualizar filas
        updated = 0
        not_found = []

        # Crear diccionario de empleados para búsqueda rápida
        emp_dict = {str(emp['employee_num']): dict(emp) for emp in employees}

        for row in range(header_row + 1, ws.max_row + 1):
            emp_num = ws.cell(row=row, column=col_map['employee_num']).value
            if emp_num and str(emp_num) in emp_dict:
                emp = emp_dict[str(emp_num)]

                if 'used' in col_map:
                    ws.cell(row=row, column=col_map['used'], value=emp['used'])
                if 'balance' in col_map:
                    ws.cell(row=row, column=col_map['balance'], value=emp['balance'])
                if 'usage_rate' in col_map:
                    ws.cell(row=row, column=col_map['usage_rate'], value=emp['usage_rate'])

                updated += 1

        # Guardar cambios
        wb.save(master_path)

        return {
            "status": "success",
            "updated_rows": updated,
            "total_employees": len(employees),
            "backup_path": backup_path,
            "columns_updated": list(col_map.keys())
        }

    except Exception as e:
        return {"status": "error", "message": str(e), "backup_path": backup_path}


def get_export_files() -> List[Dict]:
    """Lista archivos exportados disponibles"""
    files = []

    if not os.path.exists(EXPORT_DIR):
        return files

    for f in os.listdir(EXPORT_DIR):
        if f.endswith('.xlsx'):
            filepath = os.path.join(EXPORT_DIR, f)
            stat = os.stat(filepath)
            files.append({
                'filename': f,
                'path': filepath,
                'size_bytes': stat.st_size,
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
            })

    return sorted(files, key=lambda x: x['created'], reverse=True)


def cleanup_old_exports(days_to_keep: int = 30) -> Dict:
    """
    Elimina exportaciones antiguas.

    Args:
        days_to_keep: Días de antigüedad para mantener

    Returns:
        Dict con estadísticas de limpieza
    """
    if not os.path.exists(EXPORT_DIR):
        return {"deleted": 0}

    cutoff = datetime.now().timestamp() - (days_to_keep * 24 * 60 * 60)
    deleted = 0

    for f in os.listdir(EXPORT_DIR):
        filepath = os.path.join(EXPORT_DIR, f)
        if os.path.getctime(filepath) < cutoff:
            os.remove(filepath)
            deleted += 1

    return {"deleted": deleted, "days_threshold": days_to_keep}
