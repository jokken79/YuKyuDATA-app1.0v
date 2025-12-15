"""
Data Parser Agent - Agente Parseador de Datos
==============================================

Especializado en:
- Detección automática de headers en Excel
- Mapeo flexible de columnas (múltiples nombres)
- Validación de datos
- Detección de anomalías
- Transformación de formatos
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum

logger = logging.getLogger(__name__)


class ValidationSeverity(Enum):
    """Niveles de severidad para validaciones."""
    INFO = "info"
    WARNING = "warning"
    ERROR = "error"
    CRITICAL = "critical"


@dataclass
class ValidationIssue:
    """Representa un problema encontrado durante validación."""
    severity: ValidationSeverity
    field: str
    message: str
    row_index: Optional[int] = None
    value: Any = None
    suggestion: Optional[str] = None


@dataclass
class ParseResult:
    """Resultado del parsing de datos."""
    success: bool
    data: List[Dict] = field(default_factory=list)
    issues: List[ValidationIssue] = field(default_factory=list)
    stats: Dict = field(default_factory=dict)
    source_file: str = ""
    parsed_at: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class ValidationResult:
    """Resultado de validación de datos."""
    is_valid: bool
    issues: List[ValidationIssue] = field(default_factory=list)
    stats: Dict = field(default_factory=dict)

    @property
    def error_count(self) -> int:
        return sum(1 for i in self.issues if i.severity in [ValidationSeverity.ERROR, ValidationSeverity.CRITICAL])

    @property
    def warning_count(self) -> int:
        return sum(1 for i in self.issues if i.severity == ValidationSeverity.WARNING)


class DataParserAgent:
    """
    Agente especializado en parsing y validación de datos.

    Características:
    - Detección automática de encabezados
    - Mapeo flexible de columnas con múltiples nombres
    - Validación exhaustiva de datos
    - Detección de anomalías
    - Soporte para múltiples formatos (Excel, CSV)

    Ejemplo de uso:
    ```python
    parser = DataParserAgent()

    # Parsear Excel
    result = parser.parse_excel("/path/to/file.xlsx")

    # Validar datos
    validation = parser.validate_data(result.data)

    # Detectar anomalías
    anomalies = parser.detect_anomalies(result.data)
    ```
    """

    # Mapeos de columnas flexibles (japonés → inglés)
    COLUMN_MAPPINGS = {
        'employee_num': [
            '社員№', '社員番号', '従業員番号', '番号',
            'id', 'no', '№', 'employee_id', 'emp_no'
        ],
        'name': [
            '氏名', '名前', 'name', '社員名', 'full_name', '従業員名'
        ],
        'haken': [
            '派遣先', '所属', '部署', '現場', '勤務先',
            'factory', 'location', 'department'
        ],
        'granted': [
            '付与数', '付与日数', '付与', '総日数', '有給残日数',
            '付与数合計', 'granted_days', 'total_granted'
        ],
        'used': [
            '消化日数', '使用日数', '使用', '消化', '取得日数',
            'used_days', 'taken_days'
        ],
        'balance': [
            '期末残高', '残高', '残り', 'balance', 'remaining',
            '有給残', '残日数'
        ],
        'expired': [
            '期限切れ', '失効', '消滅', 'expired', '時効消滅'
        ],
        'year': [
            '年度', '年', 'year', '対象年度', '会計年度'
        ],
        'grant_date': [
            '付与日', '有給発生', '発生日', 'grant_date', '付与年月日'
        ],
        'status': [
            '現在状態', '状態', 'status', '在籍状態', '雇用状態'
        ]
    }

    # Patrones de header a buscar
    HEADER_PATTERNS = ['氏名', '名前', 'name', '社員', '従業員']

    def __init__(self):
        """Inicializa el agente parseador."""
        self._parse_history: List[ParseResult] = []

    def find_header_row(self, sheet, max_rows: int = 15) -> Tuple[Optional[int], Optional[List]]:
        """
        Encuentra la fila de encabezados en una hoja Excel.

        Args:
            sheet: Hoja de openpyxl
            max_rows: Máximo de filas a escanear

        Returns:
            Tupla (índice_fila, lista_headers) o (None, None) si no encuentra
        """
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
            row_str = ' '.join(str(cell or '') for cell in row).lower()

            # Buscar patrones de header
            for pattern in self.HEADER_PATTERNS:
                if pattern.lower() in row_str:
                    return row_idx + 1, list(row)  # +1 porque Excel es 1-indexed

        return None, None

    def map_columns(self, headers: List) -> Dict[str, int]:
        """
        Mapea nombres de columnas a índices.

        Args:
            headers: Lista de nombres de columnas

        Returns:
            Dict con mapeo {campo: índice}
        """
        column_map = {}

        for field_name, variations in self.COLUMN_MAPPINGS.items():
            for idx, header in enumerate(headers):
                header_str = str(header or '').lower().strip()

                for variation in variations:
                    if variation.lower() in header_str:
                        column_map[field_name] = idx
                        break

                if field_name in column_map:
                    break

        return column_map

    def parse_excel(self, file_path: str, sheet_name: str = None) -> ParseResult:
        """
        Parsea un archivo Excel con detección inteligente.

        Args:
            file_path: Ruta al archivo Excel
            sheet_name: Nombre de la hoja (opcional, usa la activa)

        Returns:
            ParseResult con datos parseados y estadísticas
        """
        from openpyxl import load_workbook

        result = ParseResult(success=False, source_file=file_path)
        issues = []

        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active

            # Encontrar headers
            header_row, headers = self.find_header_row(sheet)

            if header_row is None:
                result.issues.append(ValidationIssue(
                    severity=ValidationSeverity.CRITICAL,
                    field="headers",
                    message="No se encontraron encabezados válidos"
                ))
                return result

            # Mapear columnas
            column_map = self.map_columns(headers)

            # Parsear datos
            data = []
            skipped = 0

            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=header_row + 1,
                values_only=True
            ), start=header_row + 1):

                record = self._parse_row(row, column_map, row_idx)

                if record:
                    # Validar record básico
                    if not record.get('name') or record['name'] == 'Unknown':
                        skipped += 1
                        continue

                    data.append(record)
                else:
                    skipped += 1

            # Agregar estadísticas
            result.data = data
            result.success = True
            result.stats = {
                'total_rows': len(data) + skipped,
                'parsed_rows': len(data),
                'skipped_rows': skipped,
                'columns_mapped': list(column_map.keys()),
                'header_row': header_row
            }

            logger.info(f"✅ Parseados {len(data)} registros de {file_path}")

        except Exception as e:
            result.issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                field="file",
                message=f"Error al parsear archivo: {str(e)}"
            ))
            logger.error(f"❌ Error parseando {file_path}: {e}")

        self._parse_history.append(result)
        return result

    def _parse_row(
        self,
        row: tuple,
        column_map: Dict[str, int],
        row_idx: int
    ) -> Optional[Dict]:
        """Parsea una fila individual."""
        try:
            def get_value(field: str, default=None):
                idx = column_map.get(field)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return val if val is not None else default
                return default

            # Extraer valores básicos
            emp_num = str(get_value('employee_num', '')).strip()
            name = str(get_value('name', 'Unknown')).strip()

            # Skip si es fila vacía o encabezado duplicado
            if not name or name == 'Unknown' or name in ['氏名', '名前']:
                return None

            # Convertir valores numéricos
            granted = self._to_float(get_value('granted', 0))
            used = self._to_float(get_value('used', 0))
            balance = self._to_float(get_value('balance'))
            expired = self._to_float(get_value('expired', 0))
            year = self._to_int(get_value('year', datetime.now().year))

            # Si no hay balance, calcularlo
            if balance is None:
                balance = granted - used - expired

            # Calcular usage_rate
            usage_rate = round((used / granted) * 100, 1) if granted > 0 else 0

            return {
                'employeeNum': emp_num,
                'name': name,
                'haken': str(get_value('haken', '')).strip(),
                'granted': granted,
                'used': used,
                'balance': balance,
                'expired': expired,
                'usageRate': usage_rate,
                'year': year,
                '_row_index': row_idx
            }

        except Exception as e:
            logger.warning(f"Error parseando fila {row_idx}: {e}")
            return None

    def validate_data(self, data: List[Dict]) -> ValidationResult:
        """
        Valida una lista de registros de datos.

        Args:
            data: Lista de diccionarios con datos

        Returns:
            ValidationResult con problemas encontrados
        """
        issues = []

        # Validaciones por registro
        for idx, record in enumerate(data):
            row_issues = self._validate_record(record, idx)
            issues.extend(row_issues)

        # Validaciones globales
        issues.extend(self._validate_duplicates(data))
        issues.extend(self._validate_consistency(data))

        # Determinar si es válido (sin errores críticos)
        has_critical = any(
            i.severity in [ValidationSeverity.ERROR, ValidationSeverity.CRITICAL]
            for i in issues
        )

        return ValidationResult(
            is_valid=not has_critical,
            issues=issues,
            stats={
                'total_records': len(data),
                'issues_found': len(issues),
                'errors': sum(1 for i in issues if i.severity == ValidationSeverity.ERROR),
                'warnings': sum(1 for i in issues if i.severity == ValidationSeverity.WARNING)
            }
        )

    def _validate_record(self, record: Dict, idx: int) -> List[ValidationIssue]:
        """Valida un registro individual."""
        issues = []

        # Campos requeridos
        if not record.get('employeeNum'):
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                field='employee_num',
                message='Número de empleado vacío',
                row_index=idx
            ))

        if not record.get('name') or record['name'] == 'Unknown':
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                field='name',
                message='Nombre requerido',
                row_index=idx
            ))

        # Validar rangos numéricos
        granted = record.get('granted', 0)
        used = record.get('used', 0)
        balance = record.get('balance', 0)

        if granted < 0:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                field='granted',
                message=f'Días otorgados negativos: {granted}',
                row_index=idx,
                value=granted
            ))

        if used < 0:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                field='used',
                message=f'Días usados negativos: {used}',
                row_index=idx,
                value=used
            ))

        if used > granted and granted > 0:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                field='used',
                message=f'Días usados ({used}) excede días otorgados ({granted})',
                row_index=idx,
                suggestion='Verificar si hay carry-over de años anteriores'
            ))

        # Validar año
        year = record.get('year')
        current_year = datetime.now().year
        if year and (year < current_year - 5 or year > current_year + 1):
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                field='year',
                message=f'Año fuera de rango esperado: {year}',
                row_index=idx,
                value=year
            ))

        return issues

    def _validate_duplicates(self, data: List[Dict]) -> List[ValidationIssue]:
        """Detecta registros duplicados."""
        issues = []
        seen = {}

        for idx, record in enumerate(data):
            key = f"{record.get('employeeNum')}_{record.get('year')}"

            if key in seen:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    field='id',
                    message=f'Registro duplicado: {key}',
                    row_index=idx,
                    suggestion=f'Primer aparición en fila {seen[key]}'
                ))
            else:
                seen[key] = idx

        return issues

    def _validate_consistency(self, data: List[Dict]) -> List[ValidationIssue]:
        """Valida consistencia global de datos."""
        issues = []

        # Verificar si hay datos
        if not data:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                field='data',
                message='No hay datos para validar'
            ))

        return issues

    def detect_anomalies(self, data: List[Dict]) -> List[Dict]:
        """
        Detecta valores anómalos en los datos.

        Args:
            data: Lista de registros

        Returns:
            Lista de anomalías encontradas
        """
        anomalies = []

        # Calcular estadísticas para detección
        if not data:
            return anomalies

        granted_values = [r.get('granted', 0) for r in data if r.get('granted')]
        used_values = [r.get('used', 0) for r in data if r.get('used')]

        avg_granted = sum(granted_values) / len(granted_values) if granted_values else 0
        avg_used = sum(used_values) / len(used_values) if used_values else 0

        for idx, record in enumerate(data):
            # Anomalía: uso excesivo (>100% de otorgado)
            if record.get('used', 0) > record.get('granted', 0) * 1.5:
                anomalies.append({
                    'type': 'EXCESSIVE_USE',
                    'record_index': idx,
                    'employee': record.get('name'),
                    'message': f"Uso excesivo: {record.get('used')} de {record.get('granted')} días",
                    'severity': 'high'
                })

            # Anomalía: días otorgados muy altos
            if record.get('granted', 0) > 40:  # Max legal en Japón es ~40
                anomalies.append({
                    'type': 'HIGH_GRANTED',
                    'record_index': idx,
                    'employee': record.get('name'),
                    'message': f"Días otorgados inusuales: {record.get('granted')}",
                    'severity': 'medium'
                })

            # Anomalía: sin uso cuando hay días disponibles
            if record.get('granted', 0) >= 10 and record.get('used', 0) == 0:
                anomalies.append({
                    'type': 'NO_USE',
                    'record_index': idx,
                    'employee': record.get('name'),
                    'message': f"Sin uso de días disponibles ({record.get('granted')} días)",
                    'severity': 'low'
                })

        return anomalies

    # ========================================
    # UTILIDADES
    # ========================================

    def _to_float(self, value, default: float = 0.0) -> float:
        """Convierte valor a float de forma segura."""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _to_int(self, value, default: int = 0) -> int:
        """Convierte valor a int de forma segura."""
        if value is None:
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default

    def get_parse_history(self, limit: int = 10) -> List[ParseResult]:
        """Obtiene historial de parsing."""
        return self._parse_history[-limit:]


# Instancia global
_parser_instance: Optional[DataParserAgent] = None


def get_parser() -> DataParserAgent:
    """Obtiene la instancia global del parser."""
    global _parser_instance
    if _parser_instance is None:
        _parser_instance = DataParserAgent()
    return _parser_instance
