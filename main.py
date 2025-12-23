from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field, validator
from typing import Optional, List
from collections import defaultdict
from time import time
import uvicorn
import shutil
import os
import jwt
from pathlib import Path
from datetime import datetime, timedelta, date

# Local modules
import database
import excel_service
from logger import logger, log_api_request, log_db_operation, log_sync_event, log_leave_request
from fiscal_year import (
    process_year_end_carryover,
    get_employee_balance_breakdown,
    check_expiring_soon,
    check_5day_compliance,
    get_grant_recommendation,
    calculate_seniority_years,
    calculate_granted_days,
    get_fiscal_period,
    apply_lifo_deduction,
    FISCAL_CONFIG,
    GRANT_TABLE
)
from excel_export import (
    create_approved_requests_excel,
    create_monthly_report_excel,
    create_annual_ledger_excel,
    update_master_excel,
    get_export_files,
    cleanup_old_exports,
    EXPORT_DIR
)

# ============================================
# PYDANTIC MODELS FOR VALIDATION
# ============================================

class LeaveRequestCreate(BaseModel):
    employee_num: str = Field(..., min_length=1, description="Employee number")
    employee_name: str = Field(..., min_length=1, description="Employee name")
    start_date: str = Field(..., description="Start date YYYY-MM-DD")
    end_date: str = Field(..., description="End date YYYY-MM-DD")
    days_requested: float = Field(..., ge=0, le=40, description="Days requested")
    hours_requested: float = Field(0, ge=0, le=320, description="Hours requested")
    leave_type: str = Field(..., description="Leave type: full, half_am, half_pm, hourly")
    reason: Optional[str] = None

    @validator('leave_type')
    def validate_leave_type(cls, v):
        valid_types = ['full', 'half_am', 'half_pm', 'hourly']
        if v not in valid_types:
            raise ValueError(f'leave_type must be one of: {valid_types}')
        return v

    @validator('end_date')
    def validate_dates(cls, v, values):
        if 'start_date' in values:
            if v < values['start_date']:
                raise ValueError('end_date must be after start_date')
        return v


class DateRangeQuery(BaseModel):
    start_date: str
    end_date: str


# ============================================
# RATE LIMITER
# ============================================

class RateLimiter:
    """Simple in-memory rate limiter"""
    def __init__(self, max_requests: int = 100, window_seconds: int = 60):
        self.max_requests = max_requests
        self.window = window_seconds
        self.requests = defaultdict(list)

    def is_allowed(self, client_ip: str) -> bool:
        now = time()
        # Clean old requests
        self.requests[client_ip] = [
            t for t in self.requests[client_ip]
            if now - t < self.window
        ]

        if len(self.requests[client_ip]) >= self.max_requests:
            return False

        self.requests[client_ip].append(now)
        return True

    def get_remaining(self, client_ip: str) -> int:
        now = time()
        self.requests[client_ip] = [
            t for t in self.requests[client_ip]
            if now - t < self.window
        ]
        return max(0, self.max_requests - len(self.requests[client_ip]))


rate_limiter = RateLimiter(max_requests=100, window_seconds=60)


# ============================================
# JWT AUTHENTICATION SYSTEM
# ============================================

# JWT Configuration
JWT_SECRET_KEY = os.getenv("JWT_SECRET_KEY", "yukyu-secret-key-2024-change-in-production")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Security
security = HTTPBearer(auto_error=False)

# User database (simple in-memory for now)
USERS_DB = {
    "admin": {
        "password": "admin123",
        "role": "admin",
        "name": "Administrator"
    }
}


class UserLogin(BaseModel):
    """Login request model"""
    username: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)


class TokenResponse(BaseModel):
    """Token response model"""
    access_token: str
    token_type: str = "bearer"
    expires_in: int
    user: dict


def create_jwt_token(username: str, role: str) -> str:
    """Create a JWT token for authenticated user"""
    expiration = datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "sub": username,
        "role": role,
        "exp": expiration,
        "iat": datetime.utcnow()
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)


def verify_jwt_token(token: str) -> dict:
    """Verify and decode JWT token"""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """
    Dependency to get current authenticated user.
    Returns user info if authenticated, None if not.
    """
    if credentials is None:
        return None

    try:
        payload = verify_jwt_token(credentials.credentials)
        return {
            "username": payload.get("sub"),
            "role": payload.get("role")
        }
    except HTTPException:
        return None


async def require_auth(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """
    Dependency that requires authentication.
    Raises 401 if not authenticated.
    """
    if credentials is None:
        raise HTTPException(status_code=401, detail="Authentication required")

    payload = verify_jwt_token(credentials.credentials)
    return {
        "username": payload.get("sub"),
        "role": payload.get("role")
    }


async def require_admin(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """
    Dependency that requires admin role.
    Raises 401 if not authenticated, 403 if not admin.
    """
    if credentials is None:
        raise HTTPException(status_code=401, detail="Authentication required")

    payload = verify_jwt_token(credentials.credentials)
    role = payload.get("role")

    if role != "admin":
        raise HTTPException(status_code=403, detail="Admin privileges required")

    return {
        "username": payload.get("sub"),
        "role": role
    }


# ============================================
# APP INITIALIZATION
# ============================================

app = FastAPI(
    title="YuKyu Dashboard API",
    description="""
    ## Employee Paid Leave Management System (有給休暇管理システム)

    Sistema completo de gestión de vacaciones pagadas conforme a la normativa japonesa.

    ### Características Principales

    * **Gestión de Vacaciones**: Seguimiento de días otorgados, usados y balance
    * **Solicitudes de Ausencia**: Flujo de aprobación de solicitudes
    * **Cumplimiento Legal**: Verificación automática de 5 días obligatorios
    * **Reportes**: Generación de reportes mensuales y anuales
    * **Sincronización Excel**: Carga bidireccional desde archivos Excel
    * **Registro de Empleados**: Gestión de empleados Genzai (派遣) y Ukeoi (請負)

    ### Normativa

    Cumple con el Labor Standards Act Article 39 de Japón y reforma laboral de 2019.

    ### Autenticación

    La API utiliza JWT Bearer tokens. Usa `/api/auth/login` para obtener un token.
    """,
    version="2.0.0",
    contact={
        "name": "YuKyuDATA Support",
        "email": "support@yukyu.example.com"
    },
    license_info={
        "name": "Proprietary",
    },
    openapi_tags=[
        {"name": "Authentication", "description": "Autenticación y gestión de tokens JWT"},
        {"name": "Employees", "description": "Gestión de datos de empleados y vacaciones"},
        {"name": "Leave Requests", "description": "Solicitudes de vacaciones y aprobaciones"},
        {"name": "Compliance", "description": "Verificaciones de cumplimiento normativo"},
        {"name": "Analytics", "description": "Análisis y KPIs de uso de vacaciones"},
        {"name": "Reports", "description": "Generación de reportes Excel y PDF"},
        {"name": "Genzai", "description": "Gestión de empleados dispatch (派遣社員)"},
        {"name": "Ukeoi", "description": "Gestión de empleados contrato (請負社員)"},
        {"name": "System", "description": "Información del sistema y diagnóstico"},
    ],
    docs_url="/docs",  # Swagger UI
    redoc_url="/redoc",  # ReDoc
    openapi_url="/openapi.json"
)

# Configure CORS - Restricted to specific origins
# Get ports from environment variables (defaults for safe fallback)
SERVER_PORT = int(os.getenv("PORT", "8000"))
FRONTEND_PORT = int(os.getenv("FRONTEND_PORT", "3000"))

ALLOWED_ORIGINS = [
    f"http://localhost:{SERVER_PORT}",
    f"http://127.0.0.1:{SERVER_PORT}",
    f"http://localhost:{FRONTEND_PORT}",
    f"http://127.0.0.1:{FRONTEND_PORT}",
    # Also allow standard ports just in case
    "http://localhost:8000",
    "http://127.0.0.1:8000",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
]

logger.info(f"CORS Configured for Server Port: {SERVER_PORT}, Frontend Port: {FRONTEND_PORT}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,  # Changed to True to allow cookies/auth headers if needed
    allow_methods=["GET", "POST", "DELETE", "PUT", "PATCH", "OPTIONS"],
    allow_headers=["*"],     # Allow all headers including Authorization
)

# Constants - Relative paths from project directory
PROJECT_DIR = Path(__file__).parent  # Directorio del proyecto
DEFAULT_EXCEL_PATH = PROJECT_DIR / "有給休暇管理.xlsm"
EMPLOYEE_REGISTRY_PATH = PROJECT_DIR / "【新】社員台帳(UNS)T　2022.04.05～.xlsm"
UPLOAD_DIR = PROJECT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Initialize Database
database.init_db()

# ============================================
# AUTO-SYNC ON STARTUP IF DATABASE IS EMPTY
# ============================================
def auto_sync_on_startup():
    """
    Sincroniza automáticamente los datos desde Excel si la base de datos está vacía.
    Esto asegura que los datos persisten y no hay que sincronizar manualmente cada vez.
    También crea un backup automático si la BD tiene datos.
    """
    try:
        # Check if employees table is empty
        employees = database.get_employees()

        # If database has data, create automatic backup on startup
        if len(employees) > 0:
            try:
                backup_result = database.create_backup()
                logger.info(f"🔒 Auto-backup created: {backup_result['filename']}")
            except Exception as backup_err:
                logger.warning(f"⚠️ Auto-backup failed (non-critical): {str(backup_err)}")

        if len(employees) == 0:
            logger.info("📊 Database is empty - attempting auto-sync from Excel...")

            # Try to sync vacation data
            if DEFAULT_EXCEL_PATH.exists():
                logger.info(f"📁 Found vacation Excel: {DEFAULT_EXCEL_PATH}")
                data = excel_service.parse_excel_file(DEFAULT_EXCEL_PATH)
                database.save_employees(data)

                # Also parse usage details
                usage_details = excel_service.parse_yukyu_usage_details(DEFAULT_EXCEL_PATH)
                database.save_yukyu_usage_details(usage_details)

                logger.info(f"✅ Auto-synced {len(data)} employees + {len(usage_details)} usage details")
            else:
                logger.warning(f"⚠️ Vacation Excel not found at: {DEFAULT_EXCEL_PATH}")

            # Try to sync Genzai (dispatch employees)
            if EMPLOYEE_REGISTRY_PATH.exists():
                logger.info(f"📁 Found employee registry: {EMPLOYEE_REGISTRY_PATH}")

                genzai_data = excel_service.parse_genzai_sheet(EMPLOYEE_REGISTRY_PATH)
                database.save_genzai(genzai_data)
                logger.info(f"✅ Auto-synced {len(genzai_data)} dispatch employees (Genzai)")

                ukeoi_data = excel_service.parse_ukeoi_sheet(EMPLOYEE_REGISTRY_PATH)
                database.save_ukeoi(ukeoi_data)
                logger.info(f"✅ Auto-synced {len(ukeoi_data)} contract employees (Ukeoi)")
            else:
                logger.warning(f"⚠️ Employee registry not found at: {EMPLOYEE_REGISTRY_PATH}")
        else:
            logger.info(f"✅ Database already has {len(employees)} employees - skipping auto-sync")

    except Exception as e:
        logger.error(f"❌ Auto-sync failed: {str(e)}")
        # Don't raise - allow server to start even if sync fails

# Run auto-sync on startup
auto_sync_on_startup()

# Mount static files (css, js, etc.)
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serves the main dashboard page."""
    with open("templates/index.html", "r", encoding="utf-8") as f:
        return f.read()


# ============================================
# AUTHENTICATION ENDPOINTS
# ============================================

@app.post("/api/auth/login")
async def login(login_data: UserLogin):
    """
    Authenticate user and return JWT token.

    Credentials:
    - username: admin
    - password: admin123
    """
    username = login_data.username
    password = login_data.password

    # Check if user exists
    user = USERS_DB.get(username)
    if not user or user["password"] != password:
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password"
        )

    # Create JWT token
    token = create_jwt_token(username, user["role"])

    logger.info(f"User '{username}' logged in successfully")

    return {
        "status": "success",
        "access_token": token,
        "token_type": "bearer",
        "expires_in": JWT_EXPIRATION_HOURS * 3600,
        "user": {
            "username": username,
            "role": user["role"],
            "name": user["name"]
        }
    }


@app.get("/api/auth/me")
async def get_current_user_info(user: dict = Depends(require_auth)):
    """Get current authenticated user information."""
    user_data = USERS_DB.get(user["username"])
    return {
        "status": "success",
        "user": {
            "username": user["username"],
            "role": user["role"],
            "name": user_data["name"] if user_data else user["username"]
        }
    }


@app.post("/api/auth/logout")
async def logout(user: dict = Depends(get_current_user)):
    """
    Logout endpoint (client-side token removal).
    Server-side we just acknowledge the logout.
    """
    if user:
        logger.info(f"User '{user['username']}' logged out")
    return {"status": "success", "message": "Logged out successfully"}


@app.get("/api/auth/verify")
async def verify_token(user: dict = Depends(get_current_user)):
    """Verify if current token is valid."""
    if user:
        return {
            "status": "success",
            "valid": True,
            "user": user
        }
    return {
        "status": "success",
        "valid": False,
        "user": None
    }


# ============================================
# API ENDPOINTS
# ============================================

@app.get("/api/employees")
async def get_employees(year: int = None, enhanced: bool = False, active_only: bool = False):
    """Returns list of employees from SQLite.

    Args:
        year: Filter by year
        enhanced: If True, includes employee_type (genzai/ukeoi/staff) and employment_status
        active_only: If True, only returns employees with status '在職中'
    """
    try:
        if enhanced:
            data = database.get_employees_enhanced(year, active_only)
        else:
            data = database.get_employees(year)
        years = database.get_available_years()
        return {"data": data, "available_years": years}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sync")
async def sync_default_file():
    """Triggers auto-read of the default Excel file + individual usage dates."""
    if not DEFAULT_EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Default file not found at {DEFAULT_EXCEL_PATH}")

    try:
        # Parse summary data (totals)
        data = excel_service.parse_excel_file(DEFAULT_EXCEL_PATH)
        database.save_employees(data)

        # Parse individual usage dates (columns R-BE - v2.0 feature)
        usage_details = excel_service.parse_yukyu_usage_details(DEFAULT_EXCEL_PATH)
        database.save_yukyu_usage_details(usage_details)

        return {
            "status": "success",
            "count": len(data),
            "usage_details_count": len(usage_details),
            "message": f"Synced {len(data)} employees + {len(usage_details)} individual usage dates"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Handles manual Excel file upload and processing."""
    try:
        # Save temp file
        temp_path = UPLOAD_DIR / file.filename
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Process
        data = excel_service.parse_excel_file(str(temp_path))
        database.save_employees(data)
        
        # Cleanup
        os.remove(temp_path)
        
        return {"status": "success", "count": len(data), "message": f"Successfully imported {len(data)} records"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.delete("/api/reset")
async def reset_db(user: dict = Depends(require_admin)):
    """Reset database - requires admin authentication."""
    database.clear_database()
    logger.warning(f"Database cleared by admin: {user['username']}")
    return {"status": "success", "message": "Database cleared"}

# === GENZAI (Dispatch Employees) Endpoints ===

@app.get("/api/genzai")
async def get_genzai(status: str = None):
    """Returns list of dispatch employees from DBGenzaiX. Optional status filter."""
    try:
        data = database.get_genzai(status)
        return {"status": "success", "data": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sync-genzai")
async def sync_genzai():
    """Syncs DBGenzaiX sheet from employee registry file."""
    if not EMPLOYEE_REGISTRY_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Employee registry file not found at {EMPLOYEE_REGISTRY_PATH}")

    try:
        data = excel_service.parse_genzai_sheet(EMPLOYEE_REGISTRY_PATH)
        database.save_genzai(data)
        return {"status": "success", "count": len(data), "message": f"Genzai synced: {len(data)} dispatch employees"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Genzai sync failed: {str(e)}")

@app.delete("/api/reset-genzai")
async def reset_genzai(user: dict = Depends(require_admin)):
    """Clears genzai table - requires admin authentication."""
    database.clear_genzai()
    logger.warning(f"Genzai database cleared by admin: {user['username']}")
    return {"status": "success", "message": "Genzai database cleared"}

# === UKEOI (Contract Employees) Endpoints ===

@app.get("/api/ukeoi")
async def get_ukeoi(status: str = None):
    """Returns list of contract employees from DBUkeoiX. Optional status filter."""
    try:
        data = database.get_ukeoi(status)
        return {"status": "success", "data": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sync-ukeoi")
async def sync_ukeoi():
    """Syncs DBUkeoiX sheet from employee registry file."""
    if not EMPLOYEE_REGISTRY_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Employee registry file not found at {EMPLOYEE_REGISTRY_PATH}")

    try:
        data = excel_service.parse_ukeoi_sheet(EMPLOYEE_REGISTRY_PATH)
        database.save_ukeoi(data)
        return {"status": "success", "count": len(data), "message": f"Ukeoi synced: {len(data)} contract employees"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ukeoi sync failed: {str(e)}")

@app.delete("/api/reset-ukeoi")
async def reset_ukeoi(user: dict = Depends(require_admin)):
    """Clears ukeoi table - requires admin authentication."""
    database.clear_ukeoi()
    logger.warning(f"Ukeoi database cleared by admin: {user['username']}")
    return {"status": "success", "message": "Ukeoi database cleared"}


# === STAFF Endpoints ===

@app.get("/api/staff")
async def get_staff_employees(status: str = None, year: int = None, filter_by_year: bool = False):
    """
    Returns list of staff employees from DBStaffX.

    Args:
        status: Optional status filter (e.g., '在職中')
        year: Optional year for filtering
        filter_by_year: If True, filters by hire_date/leave_date
    """
    try:
        data = database.get_staff(status=status, year=year, active_in_year=filter_by_year)
        return {"status": "success", "data": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/sync-staff")
async def sync_staff():
    """Syncs DBStaffX sheet from employee registry file."""
    if not EMPLOYEE_REGISTRY_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Employee registry file not found at {EMPLOYEE_REGISTRY_PATH}")

    try:
        data = excel_service.parse_staff_sheet(EMPLOYEE_REGISTRY_PATH)
        database.save_staff(data)
        return {"status": "success", "count": len(data), "message": f"Staff synced: {len(data)} staff employees"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Staff sync failed: {str(e)}")


@app.delete("/api/reset-staff")
async def reset_staff(user: dict = Depends(require_admin)):
    """Clears staff table - requires admin authentication."""
    database.clear_staff()
    logger.warning(f"Staff database cleared by admin: {user['username']}")
    return {"status": "success", "message": "Staff database cleared"}


# === STATISTICS Endpoints ===

@app.get("/api/stats/by-factory")
async def get_factory_stats(year: int = None):
    """Returns vacation usage statistics grouped by factory (派遣先). Optional year filter."""
    try:
        data = database.get_stats_by_factory(year)
        return {"status": "success", "data": data, "count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === LEAVE REQUEST ENDPOINTS ===

@app.get("/api/factories")
async def get_factories(status: str = None):
    """Get unique list of factories from genzai and ukeoi tables."""
    try:
        factories = set()

        # Get factories from genzai (dispatch employees)
        genzai = database.get_genzai(status)
        for emp in genzai:
            factory = emp.get('dispatch_name')
            if factory:
                factories.add(factory)

        # Get factories from ukeoi (contract employees)
        ukeoi = database.get_ukeoi(status)
        for emp in ukeoi:
            factory = emp.get('contract_business')
            if factory:
                factories.add(factory)

        # Sort alphabetically
        factory_list = sorted(list(factories))

        return {"status": "success", "data": factory_list, "count": len(factory_list)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/employees/search")
async def search_employees(q: str = "", status: str = None, factory: str = None):
    """Search for employees in genzai and ukeoi tables. Optional status and factory filters."""
    try:
        results = []

        # Search in genzai (dispatch employees)
        genzai = database.get_genzai(status)
        for emp in genzai:
            emp_factory = emp.get('dispatch_name', '')

            # Apply factory filter if specified
            if factory and emp_factory != factory:
                continue

            # Apply search query if specified
            if q:
                if not (q.lower() in emp.get('name', '').lower() or
                        q.lower() in emp.get('employee_num', '').lower() or
                        q.lower() in emp_factory.lower()):
                    continue

            results.append({
                "employee_num": emp.get('employee_num'),
                "name": emp.get('name'),
                "factory": emp_factory,
                "status": emp.get('status'),
                "type": "派遣"
            })

        # Search in ukeoi (contract employees)
        ukeoi = database.get_ukeoi(status)
        for emp in ukeoi:
            emp_factory = emp.get('contract_business', '')

            # Apply factory filter if specified
            if factory and emp_factory != factory:
                continue

            # Apply search query if specified
            if q:
                if not (q.lower() in emp.get('name', '').lower() or
                        q.lower() in emp.get('employee_num', '').lower() or
                        q.lower() in emp_factory.lower()):
                    continue

            results.append({
                "employee_num": emp.get('employee_num'),
                "name": emp.get('name'),
                "factory": emp_factory,
                "status": emp.get('status'),
                "type": "請負"
            })

        return {"status": "success", "data": results, "count": len(results)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/employees/{employee_num}/leave-info")
async def get_employee_leave_info(employee_num: str):
    """Get complete leave information for an employee (current + previous year)."""
    try:
        from datetime import datetime
        current_year = datetime.now().year

        # Get employee data from genzai or ukeoi
        employee_data = None
        hourly_wage = 0

        genzai_list = database.get_genzai()
        for emp in genzai_list:
            if emp.get('employee_num') == employee_num:
                hourly_wage = emp.get('hourly_wage', 0)
                employee_data = {
                    "employee_num": emp.get('employee_num'),
                    "name": emp.get('name'),
                    "factory": emp.get('dispatch_name'),
                    "status": emp.get('status'),
                    "type": "派遣",
                    "hourly_wage": hourly_wage
                }
                break

        if not employee_data:
            ukeoi_list = database.get_ukeoi()
            for emp in ukeoi_list:
                if emp.get('employee_num') == employee_num:
                    hourly_wage = emp.get('hourly_wage', 0)
                    employee_data = {
                        "employee_num": emp.get('employee_num'),
                        "name": emp.get('name'),
                        "factory": emp.get('contract_business'),
                        "status": emp.get('status'),
                        "type": "請負",
                        "hourly_wage": hourly_wage
                    }
                    break

        if not employee_data:
            raise HTTPException(status_code=404, detail="Employee not found")

        # Get yukyu history (current + previous year)
        history = database.get_employee_yukyu_history(employee_num, current_year)

        # Get pending requests
        pending_requests = database.get_leave_requests(status='PENDING', employee_num=employee_num)

        # Calculate total available (sum of balances from history)
        total_available = sum(record.get('balance', 0) for record in history)

        # Calculate hours available (1 day = 8 hours)
        total_hours_available = total_available * 8

        # Get usage history (individual dates when yukyu was used)
        usage_history = []
        for year_record in history:
            year = year_record.get('year')
            if year:
                usage_details = database.get_yukyu_usage_details(year=year, employee_num=employee_num)
                for detail in usage_details:
                    usage_history.append({
                        'date': detail.get('use_date'),
                        'days': detail.get('days_used', 1),
                        'year': year
                    })

        # Also include approved requests as usage
        approved_requests = database.get_leave_requests(status='APPROVED', employee_num=employee_num)
        for req in approved_requests:
            # Check if not already in usage_history
            req_date = req.get('start_date')
            if req_date and not any(u['date'] == req_date for u in usage_history):
                usage_history.append({
                    'date': req_date,
                    'days': req.get('days_requested', 0),
                    'hours': req.get('hours_requested', 0),
                    'type': req.get('leave_type', 'full'),
                    'year': req.get('year'),
                    'source': 'request'
                })

        # Sort by date descending (newest first)
        usage_history.sort(key=lambda x: x.get('date', ''), reverse=True)

        return {
            "status": "success",
            "employee": employee_data,
            "yukyu_history": history,
            "usage_history": usage_history,  # Individual usage dates
            "total_available": round(total_available, 1),
            "total_hours_available": round(total_hours_available, 1),
            "hourly_wage": hourly_wage,
            "pending_requests": pending_requests
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/leave-requests")
async def create_leave_request(request_data: dict):
    """Create a new leave request with support for 時間単位有給 (hourly leave)."""
    try:
        from datetime import datetime

        # Validate required fields
        required = ['employee_num', 'employee_name', 'start_date', 'end_date', 'days_requested']
        for field in required:
            if field not in request_data:
                raise HTTPException(status_code=400, detail=f"Missing required field: {field}")

        # Get current year
        current_year = datetime.now().year

        # Validate employee has sufficient balance
        history = database.get_employee_yukyu_history(request_data['employee_num'], current_year)
        total_available = sum(record.get('balance', 0) for record in history)

        # Convert hours to days for validation (8 hours = 1 day)
        hours_requested = request_data.get('hours_requested', 0)
        total_days_equivalent = request_data['days_requested'] + (hours_requested / 8)

        if total_days_equivalent > total_available:
            raise HTTPException(
                status_code=400,
                detail=f"残日数が不足しています。残り: {total_available}日, 申請: {total_days_equivalent}日相当"
            )

        # Get hourly wage from genzai or ukeoi
        hourly_wage = 0
        genzai_list = database.get_genzai()
        for emp in genzai_list:
            if emp.get('employee_num') == request_data['employee_num']:
                hourly_wage = emp.get('hourly_wage', 0)
                break

        if hourly_wage == 0:
            ukeoi_list = database.get_ukeoi()
            for emp in ukeoi_list:
                if emp.get('employee_num') == request_data['employee_num']:
                    hourly_wage = emp.get('hourly_wage', 0)
                    break

        # Create request with new fields
        request_id = database.create_leave_request(
            employee_num=request_data['employee_num'],
            employee_name=request_data['employee_name'],
            start_date=request_data['start_date'],
            end_date=request_data['end_date'],
            days_requested=request_data['days_requested'],
            reason=request_data.get('reason', ''),
            year=current_year,
            hours_requested=hours_requested,
            leave_type=request_data.get('leave_type', 'full'),
            hourly_wage=hourly_wage
        )

        return {
            "status": "success",
            "message": "申請が作成されました",
            "request_id": request_id,
            "hourly_wage": hourly_wage,
            "cost_estimate": ((request_data['days_requested'] * 8) + hours_requested) * hourly_wage
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/leave-requests")
async def get_leave_requests_list(status: str = None, employee_num: str = None, year: int = None):
    """Get list of leave requests with optional filters."""
    try:
        requests = database.get_leave_requests(status=status, employee_num=employee_num, year=year)
        return {"status": "success", "data": requests, "count": len(requests)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/leave-requests/{request_id}/approve")
async def approve_leave_request(request_id: int, approval_data: dict):
    """Approve a leave request and automatically update yukyu balance."""
    try:
        approved_by = approval_data.get('approved_by', 'Manager')

        # Approve request (this also updates the yukyu balance automatically)
        database.approve_leave_request(request_id, approved_by)

        return {
            "status": "success",
            "message": "Request approved and yukyu balance updated"
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/leave-requests/{request_id}/reject")
async def reject_leave_request(request_id: int, rejection_data: dict):
    """Reject a leave request."""
    try:
        rejected_by = rejection_data.get('rejected_by', 'Manager')

        database.reject_leave_request(request_id, rejected_by)

        return {
            "status": "success",
            "message": "Request rejected"
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/leave-requests/{request_id}")
async def cancel_leave_request(request_id: int, user: dict = Depends(require_auth)):
    """
    Cancela una solicitud PENDIENTE - requires authentication.
    Solo funciona si el status es 'PENDING'.
    La solicitud se elimina completamente.
    """
    try:
        result = database.cancel_leave_request(request_id)
        logger.info(f"Leave request {request_id} cancelled by {user['username']}: {result}")

        return {
            "status": "success",
            "message": f"申請 #{request_id} がキャンセルされました",
            "cancelled": result
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Cancel request error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/leave-requests/{request_id}/revert")
async def revert_leave_request(request_id: int, revert_data: dict = None):
    """
    Revierte una solicitud YA APROBADA.
    Devuelve los días usados al balance del empleado.
    El status cambia a 'CANCELLED'.
    """
    try:
        if revert_data is None:
            revert_data = {}

        reverted_by = revert_data.get('reverted_by', 'Manager')
        result = database.revert_approved_request(request_id, reverted_by)
        logger.info(f"Leave request {request_id} reverted: {result}")

        return {
            "status": "success",
            "message": f"申請 #{request_id} が取り消されました。{result['days_returned']}日が返却されました",
            "reverted": result
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Revert request error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# === BACKUP & RESTORE ENDPOINTS ===

@app.post("/api/backup")
async def create_backup():
    """
    Crea una copia de seguridad de la base de datos.
    Los backups se guardan en la carpeta 'backups/'.
    Solo se mantienen los últimos 10 backups.
    """
    try:
        result = database.create_backup()
        logger.info(f"Backup created: {result['filename']}")

        return {
            "status": "success",
            "message": f"Backup creado: {result['filename']}",
            "backup": result
        }
    except Exception as e:
        logger.error(f"Backup error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/backups")
async def list_backups():
    """Lista todos los backups disponibles."""
    try:
        backups = database.list_backups()
        return {
            "status": "success",
            "count": len(backups),
            "backups": backups
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/backup/restore")
async def restore_backup(restore_data: dict):
    """
    Restaura la base de datos desde un backup.
    CUIDADO: Esto sobrescribe la base de datos actual.
    Se crea un backup automático antes de restaurar.

    Body: {"filename": "yukyu_backup_20250115_123456.db"}
    """
    try:
        filename = restore_data.get('filename')
        if not filename:
            raise HTTPException(status_code=400, detail="filename is required")

        result = database.restore_backup(filename)
        logger.info(f"Backup restored: {result}")

        return {
            "status": "success",
            "message": f"Base de datos restaurada desde {filename}",
            "restore": result
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Restore error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# === YUKYU USAGE DETAILS ENDPOINTS (v2.0 Features) ===

@app.get("/api/yukyu/usage-details")
async def get_usage_details(employee_num: str = None, year: int = None, month: int = None):
    """Get individual yukyu usage dates (v2.0 feature: 使用日一覧)."""
    try:
        details = database.get_yukyu_usage_details(employee_num, year, month)
        return {"status": "success", "data": details, "count": len(details)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/yukyu/monthly-summary/{year}")
async def get_monthly_summary(year: int):
    """Get monthly usage summary for a year (v2.0 feature: 月別フィルター)."""
    try:
        summary = database.get_monthly_usage_summary(year)

        # Also get employee list for each month
        monthly_data = []
        for month in range(1, 13):
            month_details = database.get_yukyu_usage_details(year=year, month=month)

            # Get unique employees for this month
            employees_in_month = {}
            for detail in month_details:
                emp_num = detail['employee_num']
                if emp_num not in employees_in_month:
                    employees_in_month[emp_num] = {
                        'employee_num': emp_num,
                        'name': detail['name'],
                        'days_used': 0,
                        'dates': []
                    }
                employees_in_month[emp_num]['days_used'] += detail['days_used']
                employees_in_month[emp_num]['dates'].append(detail['use_date'])

            monthly_data.append({
                'month': month,
                'employee_count': len(employees_in_month),
                'total_days': summary.get(month, {}).get('total_days', 0),
                'employees': list(employees_in_month.values())
            })

        return {"status": "success", "year": year, "data": monthly_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/yukyu/kpi-stats/{year}")
async def get_kpi_stats(year: int):
    """Get correct KPI stats based on individual usage dates (R-BE columns).

    This returns the TRUE usage total from individual dates,
    not the column N sum which represents grant-period totals.
    """
    try:
        # Get employees for count and granted totals
        employees = database.get_employees(year=year)
        total_employees = len(employees)
        total_granted = sum(emp.get('granted', 0) for emp in employees)

        # Get TRUE usage from individual dates (R-BE columns)
        usage_details = database.get_yukyu_usage_details(year=year)
        total_used = sum(detail.get('days_used', 0) for detail in usage_details)

        # Calculate balance and rate
        total_balance = total_granted - total_used
        usage_rate = round((total_used / total_granted) * 100) if total_granted > 0 else 0

        return {
            "status": "success",
            "year": year,
            "total_employees": total_employees,
            "total_granted": round(total_granted, 1),
            "total_used": round(total_used, 1),
            "total_balance": round(total_balance, 1),
            "usage_rate": usage_rate
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/yukyu/by-employee-type/{year}")
async def get_usage_by_employee_type(year: int):
    """Get yukyu usage breakdown by employee type (派遣/請負/スタッフ) for a year."""
    try:
        # Get all yukyu data for the year
        yukyu_data = database.get_employees(year=year)

        # Get employee lists
        genzai_employees = database.get_genzai()
        ukeoi_employees = database.get_ukeoi()

        # Create sets for quick lookup
        genzai_nums = {emp['employee_num'] for emp in genzai_employees if emp.get('employee_num')}
        ukeoi_nums = {emp['employee_num'] for emp in ukeoi_employees if emp.get('employee_num')}

        # Classify and aggregate
        hakenshain = {'employees': [], 'total_used': 0, 'count': 0}
        ukeoi = {'employees': [], 'total_used': 0, 'count': 0}
        staff = {'employees': [], 'total_used': 0, 'count': 0}

        for emp in yukyu_data:
            emp_num = str(emp['employee_num']) if emp['employee_num'] else None

            if emp_num in genzai_nums:
                hakenshain['employees'].append(emp)
                hakenshain['total_used'] += emp['used']
                hakenshain['count'] += 1
            elif emp_num in ukeoi_nums:
                ukeoi['employees'].append(emp)
                ukeoi['total_used'] += emp['used']
                ukeoi['count'] += 1
            elif emp_num:
                staff['employees'].append(emp)
                staff['total_used'] += emp['used']
                staff['count'] += 1

        total_used = hakenshain['total_used'] + ukeoi['total_used'] + staff['total_used']

        return {
            "status": "success",
            "year": year,
            "total_used": total_used,
            "breakdown": {
                "hakenshain": {
                    "count": hakenshain['count'],
                    "total_used": hakenshain['total_used'],
                    "percentage": round((hakenshain['total_used'] / total_used * 100), 1) if total_used > 0 else 0
                },
                "ukeoi": {
                    "count": ukeoi['count'],
                    "total_used": ukeoi['total_used'],
                    "percentage": round((ukeoi['total_used'] / total_used * 100), 1) if total_used > 0 else 0
                },
                "staff": {
                    "count": staff['count'],
                    "total_used": staff['total_used'],
                    "percentage": round((staff['total_used'] / total_used * 100), 1) if total_used > 0 else 0
                }
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === COMPLIANCE & AGENT ENDPOINTS ===

@app.get("/api/compliance/5day-check/{year}")
async def check_5day_compliance(year: int):
    """
    Verifica cumplimiento de 5日取得義務 para todos los empleados.
    Retorna empleados que no cumplen con el mínimo de 5 días.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()
        results = compliance.check_all_5_day_compliance(year)
        return {
            "status": "success",
            "year": year,
            "summary": {
                "total_checked": results['total_checked'],
                "compliant": results['compliant'],
                "at_risk": results['at_risk'],
                "non_compliant": results['non_compliant'],
                "compliance_rate": results.get('compliance_rate', 0)
            },
            "non_compliant_employees": [
                {
                    "employee_num": c.employee_num,
                    "name": c.employee_name,
                    "days_used": c.days_used,
                    "days_remaining": c.days_remaining,
                    "status": c.status.value
                }
                for c in results['checks']
                if c.status.value != 'compliant'
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/compliance/expiring/{year}")
async def check_expiring_balances(year: int, warning_days: int = 30):
    """
    Verifica balances próximos a expirar.
    Retorna empleados con días que expiran dentro del período de alerta.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()
        results = compliance.check_expiring_balances(year, warning_days)
        return {
            "status": "success",
            "year": year,
            "warning_days": warning_days,
            "count": len(results),
            "employees": [
                {
                    "employee_num": r.employee_num,
                    "name": r.employee_name,
                    "expiring_days": r.expiring_days,
                    "expiry_date": r.expiry_date,
                    "days_until_expiry": r.days_until_expiry,
                    "alert_level": r.alert_level.value
                }
                for r in results
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/compliance/report/{year}")
async def get_compliance_report(year: int):
    """
    Genera un reporte completo de compliance para el año especificado.
    Incluye: 5日取得義務, expiración, alertas.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()
        report = compliance.get_compliance_report(year)
        return {"status": "success", "report": report}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/compliance/alerts")
async def get_compliance_alerts():
    """
    Obtiene todas las alertas de compliance activas.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()
        alerts = compliance.get_active_alerts()
        summary = compliance.get_alerts_summary()
        return {
            "status": "success",
            "summary": summary,
            "alerts": [
                {
                    "id": a.alert_id,
                    "level": a.level.value,
                    "type": a.type,
                    "employee_num": a.employee_num,
                    "employee_name": a.employee_name,
                    "message": a.message,
                    "message_ja": a.message_ja,
                    "action_required": a.action_required,
                    "created_at": a.created_at
                }
                for a in alerts
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/compliance/annual-ledger/{year}")
async def get_annual_ledger(year: int):
    """
    Genera el 年次有給休暇管理簿 (libro de gestión de vacaciones anuales).
    Documento requerido por ley desde 2019.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()
        entries = compliance.generate_annual_ledger(year)
        return {
            "status": "success",
            "year": year,
            "document_name": "年次有給休暇管理簿",
            "count": len(entries),
            "entries": [
                {
                    "employee_num": e.employee_num,
                    "employee_name": e.employee_name,
                    "grant_date": e.grant_date,
                    "granted_days": e.granted_days,
                    "used_dates": e.used_dates,
                    "used_days": e.used_days,
                    "remaining_days": e.remaining_days
                }
                for e in entries
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/compliance/export-ledger/{year}")
async def export_annual_ledger(year: int, format: str = "csv"):
    """
    Exporta el 年次有給休暇管理簿 a archivo.
    Formatos: csv, json
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()

        filename = f"年次有給休暇管理簿_{year}.{format}"
        output_path = UPLOAD_DIR / filename

        success = compliance.export_annual_ledger(year, str(output_path), format)

        if success:
            return {
                "status": "success",
                "message": f"年次有給休暇管理簿 exported successfully",
                "filename": filename,
                "path": str(output_path)
            }
        else:
            raise HTTPException(status_code=500, detail="Export failed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === ACTIVE EMPLOYEES FILTER (Solo empleados activos) ===

def get_active_employee_nums() -> set:
    """
    Obtiene los números de empleados ACTIVOS (status = '在職中').
    Cruza genzai y ukeoi para obtener solo los que están trabajando.
    """
    active_nums = set()

    # Empleados activos de genzai (派遣)
    genzai = database.get_genzai(status='在職中')
    for emp in genzai:
        if emp.get('employee_num'):
            active_nums.add(str(emp['employee_num']))

    # Empleados activos de ukeoi (請負)
    ukeoi = database.get_ukeoi(status='在職中')
    for emp in ukeoi:
        if emp.get('employee_num'):
            active_nums.add(str(emp['employee_num']))

    return active_nums


@app.get("/api/employees/active")
async def get_active_employees(year: int = None):
    """
    Obtiene solo empleados ACTIVOS (在職中) con sus datos de yukyu.
    Cruza employees con genzai/ukeoi para filtrar renunciados.
    """
    try:
        active_nums = get_active_employee_nums()

        # Obtener datos de yukyu
        employees = database.get_employees(year=year)

        # Filtrar solo activos
        active_employees = [
            emp for emp in employees
            if str(emp.get('employee_num', '')) in active_nums
        ]

        years = database.get_available_years()

        return {
            "status": "success",
            "data": active_employees,
            "count": len(active_employees),
            "total_in_db": len(employees),
            "filtered_out": len(employees) - len(active_employees),
            "available_years": years
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/employees/by-type")
async def get_employees_by_type(year: int = None, active_only: bool = True, filter_by_year: bool = True):
    """
    Obtiene empleados separados por tipo: Haken (派遣), Ukeoi (請負), Staff.

    Args:
        year: Año fiscal para filtrar datos de vacaciones
        active_only: Si True, solo muestra empleados con status '在職中'
        filter_by_year: Si True, filtra empleados que estaban activos durante ese año
                       (入社日 <= año AND (退社日 IS NULL OR 退社日 >= año))
    """
    try:
        employees = database.get_employees(year=year)

        # Obtener listas de empleados por tipo con filtros
        genzai_list = database.get_genzai(
            status='在職中' if active_only else None,
            year=year if filter_by_year else None,
            active_in_year=filter_by_year
        )
        ukeoi_list = database.get_ukeoi(
            status='在職中' if active_only else None,
            year=year if filter_by_year else None,
            active_in_year=filter_by_year
        )
        staff_list = database.get_staff(
            status='在職中' if active_only else None,
            year=year if filter_by_year else None,
            active_in_year=filter_by_year
        )

        genzai_nums = {str(emp['employee_num']) for emp in genzai_list if emp.get('employee_num')}
        ukeoi_nums = {str(emp['employee_num']) for emp in ukeoi_list if emp.get('employee_num')}
        staff_nums = {str(emp['employee_num']) for emp in staff_list if emp.get('employee_num')}

        # Clasificar empleados
        haken_employees = []
        ukeoi_employees = []
        staff_employees = []

        for emp in employees:
            emp_num = str(emp.get('employee_num', ''))

            # Enriquecer con datos adicionales
            emp_enriched = dict(emp)

            if emp_num in genzai_nums:
                # Buscar datos adicionales de genzai
                genzai_data = next((g for g in genzai_list if str(g.get('employee_num')) == emp_num), {})
                emp_enriched['type'] = 'haken'
                emp_enriched['dispatch_name'] = genzai_data.get('dispatch_name', '')
                emp_enriched['status'] = genzai_data.get('status', '')
                emp_enriched['hourly_wage'] = genzai_data.get('hourly_wage', 0)
                emp_enriched['hire_date'] = genzai_data.get('hire_date', '')
                emp_enriched['leave_date'] = genzai_data.get('leave_date', '')
                haken_employees.append(emp_enriched)

            elif emp_num in ukeoi_nums:
                # Buscar datos adicionales de ukeoi
                ukeoi_data = next((u for u in ukeoi_list if str(u.get('employee_num')) == emp_num), {})
                emp_enriched['type'] = 'ukeoi'
                emp_enriched['contract_business'] = ukeoi_data.get('contract_business', '')
                emp_enriched['status'] = ukeoi_data.get('status', '')
                emp_enriched['hourly_wage'] = ukeoi_data.get('hourly_wage', 0)
                emp_enriched['hire_date'] = ukeoi_data.get('hire_date', '')
                emp_enriched['leave_date'] = ukeoi_data.get('leave_date', '')
                ukeoi_employees.append(emp_enriched)

            elif emp_num in staff_nums:
                # Buscar datos adicionales de staff
                staff_data = next((s for s in staff_list if str(s.get('employee_num')) == emp_num), {})
                emp_enriched['type'] = 'staff'
                emp_enriched['office'] = staff_data.get('office', '')
                emp_enriched['status'] = staff_data.get('status', '')
                emp_enriched['hire_date'] = staff_data.get('hire_date', '')
                emp_enriched['leave_date'] = staff_data.get('leave_date', '')
                staff_employees.append(emp_enriched)

            elif not active_only:  # Incluir todos si no hay filtro activo
                emp_enriched['type'] = 'unknown'
                staff_employees.append(emp_enriched)

        return {
            "status": "success",
            "year": year,
            "active_only": active_only,
            "haken": {
                "count": len(haken_employees),
                "employees": haken_employees,
                "total_used": sum(e.get('used', 0) for e in haken_employees),
                "total_granted": sum(e.get('granted', 0) for e in haken_employees)
            },
            "ukeoi": {
                "count": len(ukeoi_employees),
                "employees": ukeoi_employees,
                "total_used": sum(e.get('used', 0) for e in ukeoi_employees),
                "total_granted": sum(e.get('granted', 0) for e in ukeoi_employees)
            },
            "staff": {
                "count": len(staff_employees),
                "employees": staff_employees,
                "total_used": sum(e.get('used', 0) for e in staff_employees),
                "total_granted": sum(e.get('granted', 0) for e in staff_employees)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/analytics/top10-active/{year}")
async def get_top10_active_users(year: int):
    """
    Obtiene el Top 10 de usuarios activos (solo 在職中).
    Excluye empleados que ya renunciaron.
    """
    try:
        active_nums = get_active_employee_nums()
        employees = database.get_employees(year=year)

        # Filtrar solo activos y ordenar por uso
        active_employees = [
            emp for emp in employees
            if str(emp.get('employee_num', '')) in active_nums
        ]

        # Ordenar por días usados (descendente)
        sorted_active = sorted(active_employees, key=lambda x: x.get('used', 0), reverse=True)
        top10 = sorted_active[:10]

        return {
            "status": "success",
            "year": year,
            "count": len(top10),
            "data": [
                {
                    "rank": i + 1,
                    "employee_num": emp.get('employee_num'),
                    "name": emp.get('name'),
                    "haken": emp.get('haken'),
                    "used": emp.get('used', 0),
                    "granted": emp.get('granted', 0),
                    "balance": emp.get('balance', 0),
                    "usage_rate": emp.get('usage_rate', 0)
                }
                for i, emp in enumerate(top10)
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/analytics/high-balance-active/{year}")
async def get_high_balance_active(year: int, limit: int = 10):
    """
    Obtiene empleados activos con balance alto (días sin usar).
    Solo incluye empleados en estado 在職中.
    """
    try:
        active_nums = get_active_employee_nums()
        employees = database.get_employees(year=year)

        # Filtrar solo activos
        active_employees = [
            emp for emp in employees
            if str(emp.get('employee_num', '')) in active_nums
        ]

        # Ordenar por balance (descendente)
        sorted_by_balance = sorted(active_employees, key=lambda x: x.get('balance', 0), reverse=True)
        high_balance = sorted_by_balance[:limit]

        return {
            "status": "success",
            "year": year,
            "count": len(high_balance),
            "data": [
                {
                    "employee_num": emp.get('employee_num'),
                    "name": emp.get('name'),
                    "haken": emp.get('haken'),
                    "balance": emp.get('balance', 0),
                    "granted": emp.get('granted', 0),
                    "used": emp.get('used', 0)
                }
                for emp in high_balance
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# === ORCHESTRATOR ENDPOINTS (Sistema de Agentes) ===

@app.get("/api/orchestrator/status")
async def get_orchestrator_status():
    """
    Obtiene el estado actual del orquestador.
    """
    try:
        from agents.orchestrator import get_orchestrator
        orchestrator = get_orchestrator()
        status = orchestrator.get_current_status()

        return {
            "status": "success",
            "orchestrator": {
                "current_pipeline": status.get('current_pipeline'),
                "total_pipelines_executed": status.get('total_pipelines_executed', 0),
                "last_execution": status.get('last_execution')
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/orchestrator/history")
async def get_orchestrator_history(limit: int = 10):
    """
    Obtiene el historial de pipelines ejecutados.
    """
    try:
        from agents.orchestrator import get_orchestrator
        orchestrator = get_orchestrator()
        history = orchestrator.get_pipeline_history(limit)

        return {
            "status": "success",
            "count": len(history),
            "history": [
                {
                    "pipeline_name": p.pipeline_name,
                    "status": p.status.value,
                    "total_duration_ms": p.total_duration_ms,
                    "tasks_count": len(p.tasks),
                    "started_at": p.started_at,
                    "completed_at": p.completed_at
                }
                for p in history
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/orchestrator/run-compliance-check/{year}")
async def run_compliance_pipeline(year: int):
    """
    Ejecuta el pipeline completo de compliance check.
    """
    try:
        from agents.orchestrator import get_orchestrator
        from agents.compliance import get_compliance

        orchestrator = get_orchestrator()
        orchestrator.compliance_agent = get_compliance()

        result = orchestrator.orchestrate_compliance_check(year)

        return {
            "status": "success",
            "pipeline_name": result.pipeline_name,
            "pipeline_status": result.status.value,
            "duration_ms": result.total_duration_ms,
            "tasks": [
                {
                    "name": t.task_name,
                    "status": t.status.value,
                    "duration_ms": t.duration_ms,
                    "error": t.error
                }
                for t in result.tasks
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# === SYSTEM STATUS & AUDIT ENDPOINTS ===

@app.get("/api/system/snapshot")
async def get_system_snapshot():
    """
    Obtiene un snapshot del estado actual del sistema.
    Útil para monitoreo y debugging.
    """
    try:
        from agents.documentor import get_documentor
        documentor = get_documentor()
        snapshot = documentor.get_system_snapshot()
        return {"status": "success", "snapshot": snapshot.to_dict()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/audit-log")
async def get_audit_log(
    action: str = None,
    entity_type: str = None,
    limit: int = 100
):
    """
    Obtiene el historial de auditoría del sistema.
    Filtrable por tipo de acción y entidad.
    """
    try:
        from agents.documentor import get_documentor
        documentor = get_documentor()
        entries = documentor.search_history(
            action=action,
            entity_type=entity_type,
            limit=limit
        )
        return {"status": "success", "count": len(entries), "entries": entries}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/activity-report")
async def get_activity_report(days: int = 7):
    """
    Genera un reporte de actividad para los últimos N días.
    """
    try:
        from agents.documentor import get_documentor
        from datetime import timedelta
        documentor = get_documentor()

        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        report = documentor.generate_activity_report(
            start_date.isoformat(),
            end_date.isoformat()
        )
        return {"status": "success", "report": report.to_dict()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === NOTIFICATIONS ENDPOINTS ===

@app.get("/api/notifications")
async def get_notifications(employee_num: str = None, unread_only: bool = False):
    """
    Obtiene notificaciones del sistema.
    Las notificaciones incluyen alertas de compliance, solicitudes aprobadas/rechazadas, etc.
    """
    try:
        from agents.compliance import get_compliance
        compliance = get_compliance()

        # Obtener alertas como notificaciones
        alerts = compliance.get_active_alerts()

        notifications = []
        for alert in alerts:
            if employee_num and alert.employee_num != employee_num:
                continue

            notifications.append({
                "id": alert.alert_id,
                "type": alert.type,
                "level": alert.level.value,
                "title": alert.type.replace('_', ' ').title(),
                "message": alert.message_ja,
                "employee_num": alert.employee_num,
                "created_at": alert.created_at,
                "is_read": False  # Por implementar
            })

        return {
            "status": "success",
            "count": len(notifications),
            "notifications": notifications
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === CALENDAR ENDPOINTS ===

@app.get("/api/calendar/events")
async def get_calendar_events(year: int = None, month: int = None, source: str = 'requests', active_only: bool = True):
    """
    カレンダー用のイベントデータを取得。
    承認済み休暇申請のみを返す（デフォルト）。

    Args:
        year: 年度フィルター
        month: 月フィルター
        source: データソース
            - 'requests': 承認済み申請のみ (デフォルト・推奨)
            - 'excel': Excel使用詳細のみ
            - 'all': 両方（重複の可能性あり）
        active_only: 在職中の従業員のみ表示 (デフォルト: True)
    """
    try:
        if not year:
            year = datetime.now().year

        # 在職中のみフィルタリングする場合、アクティブな従業員番号を取得
        active_nums = get_active_employee_nums() if active_only else None

        events = []
        filtered_count = 0

        # 休暇タイプに応じた色分け
        type_colors = {
            'full': '#38bdf8',      # 全日休暇 - 青
            'half_am': '#818cf8',   # 午前半休 - 紫
            'half_pm': '#f472b6',   # 午後半休 - ピンク
            'hourly': '#fbbf24'     # 時間休 - 黄色
        }
        type_labels = {
            'full': '全日',
            'half_am': '午前半休',
            'half_pm': '午後半休',
            'hourly': '時間休'
        }

        # 承認済み休暇申請を取得 (source = 'requests' or 'all')
        if source in ['requests', 'all']:
            approved_requests = database.get_leave_requests(status='APPROVED', year=year)
            for req in approved_requests:
                emp_num = str(req.get('employee_num', ''))

                # フィルタリング: 在職中のみ
                if active_only and active_nums and emp_num not in active_nums:
                    filtered_count += 1
                    continue

                leave_type = req.get('leave_type', 'full')
                events.append({
                    'id': f"request_{req['id']}",
                    'title': f"{req['employee_name']} ({type_labels.get(leave_type, '休暇')})",
                    'start': req['start_date'],
                    'end': req['end_date'],
                    'color': type_colors.get(leave_type, '#38bdf8'),
                    'type': 'approved_request',
                    'employee_num': req['employee_num'],
                    'employee_name': req['employee_name'],
                    'leave_type': leave_type,
                    'days': req.get('days_requested', 0),
                    'hours': req.get('hours_requested', 0)
                })

        # 使用日詳細を取得 (source = 'excel' or 'all')
        # 注意: Excelデータは検証されていないため、不正確な場合があります
        if source in ['excel', 'all']:
            usage_details = database.get_yukyu_usage_details(year=year, month=month)
            for detail in usage_details:
                emp_num = str(detail.get('employee_num', ''))

                # フィルタリング: 在職中のみ
                if active_only and active_nums and emp_num not in active_nums:
                    filtered_count += 1
                    continue

                events.append({
                    'id': f"usage_{detail.get('id', '')}",
                    'title': f"{detail['name']} ({detail.get('days_used', 1)}日)",
                    'start': detail['use_date'],
                    'end': detail['use_date'],
                    'color': '#34d399',  # 緑
                    'type': 'usage_detail',
                    'employee_num': detail['employee_num'],
                    'employee_name': detail['name'],
                    'days': detail.get('days_used', 1)
                })

        return {
            "status": "success",
            "year": year,
            "month": month,
            "source": source,
            "active_only": active_only,
            "count": len(events),
            "filtered_out": filtered_count,
            "events": events
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/calendar/summary/{year}/{month}")
async def get_calendar_month_summary(year: int, month: int, source: str = 'requests'):
    """
    月別カレンダーサマリーを取得。
    各日の休暇取得人数を返す。

    Args:
        source: データソース ('requests', 'excel', 'all') - デフォルトは 'requests'
    """
    try:
        import calendar
        from collections import defaultdict

        # 月の日数を取得
        _, days_in_month = calendar.monthrange(year, month)

        # 日ごとの集計
        daily_counts = defaultdict(lambda: {'count': 0, 'employees': []})

        # 承認済み申請 (source = 'requests' or 'all')
        if source in ['requests', 'all']:
            approved = database.get_leave_requests(status='APPROVED', year=year)
            for req in approved:
                start = datetime.strptime(req['start_date'], '%Y-%m-%d')
                end = datetime.strptime(req['end_date'], '%Y-%m-%d')

                current = start
                while current <= end:
                    if current.year == year and current.month == month:
                        day_key = current.strftime('%Y-%m-%d')
                        daily_counts[day_key]['count'] += 1
                        daily_counts[day_key]['employees'].append({
                            'name': req['employee_name'],
                            'type': req.get('leave_type', 'full')
                        })
                    current = current + timedelta(days=1)

        # 使用日詳細 (source = 'excel' or 'all')
        if source in ['excel', 'all']:
            usage = database.get_yukyu_usage_details(year=year, month=month)
            for detail in usage:
                day_key = detail['use_date']
                # 重複チェック
                exists = any(e['name'] == detail['name'] for e in daily_counts[day_key]['employees'])
                if not exists:
                    daily_counts[day_key]['count'] += 1
                    daily_counts[day_key]['employees'].append({
                        'name': detail['name'],
                        'type': 'usage'
                    })

        return {
            "status": "success",
            "year": year,
            "month": month,
            "source": source,
            "days_in_month": days_in_month,
            "daily_summary": dict(daily_counts)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === EXPORT ENDPOINTS ===

@app.post("/api/export/excel")
async def export_to_excel(export_type: str = "employees", year: int = None):
    """
    データをExcel形式でエクスポート。
    export_type: employees, requests, compliance, calendar
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not year:
            year = datetime.now().year

        wb = Workbook()
        ws = wb.active

        # ヘッダースタイル
        header_fill = PatternFill(start_color="38bdf8", end_color="38bdf8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if export_type == "employees":
            ws.title = f"有給休暇一覧_{year}"
            headers = ["社員番号", "氏名", "派遣先", "付与日数", "使用日数", "残日数", "消化率", "年度"]

            data = database.get_employees(year=year)

            # ヘッダー
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # データ
            for row, emp in enumerate(data, 2):
                ws.cell(row=row, column=1, value=emp['employee_num']).border = thin_border
                ws.cell(row=row, column=2, value=emp['name']).border = thin_border
                ws.cell(row=row, column=3, value=emp.get('haken', '')).border = thin_border
                ws.cell(row=row, column=4, value=emp['granted']).border = thin_border
                ws.cell(row=row, column=5, value=emp['used']).border = thin_border
                ws.cell(row=row, column=6, value=emp['balance']).border = thin_border
                ws.cell(row=row, column=7, value=f"{emp.get('usage_rate', 0):.1f}%").border = thin_border
                ws.cell(row=row, column=8, value=emp['year']).border = thin_border

        elif export_type == "requests":
            ws.title = f"休暇申請一覧_{year}"
            headers = ["ID", "社員番号", "氏名", "開始日", "終了日", "種類", "日数", "時間", "理由", "ステータス", "申請日"]

            data = database.get_leave_requests(year=year)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            type_labels = {'full': '全日', 'half_am': '午前半休', 'half_pm': '午後半休', 'hourly': '時間休'}
            status_labels = {'PENDING': '審査中', 'APPROVED': '承認済', 'REJECTED': '却下'}

            for row, req in enumerate(data, 2):
                ws.cell(row=row, column=1, value=req['id']).border = thin_border
                ws.cell(row=row, column=2, value=req['employee_num']).border = thin_border
                ws.cell(row=row, column=3, value=req['employee_name']).border = thin_border
                ws.cell(row=row, column=4, value=req['start_date']).border = thin_border
                ws.cell(row=row, column=5, value=req['end_date']).border = thin_border
                ws.cell(row=row, column=6, value=type_labels.get(req.get('leave_type', 'full'), '')).border = thin_border
                ws.cell(row=row, column=7, value=req.get('days_requested', 0)).border = thin_border
                ws.cell(row=row, column=8, value=req.get('hours_requested', 0)).border = thin_border
                ws.cell(row=row, column=9, value=req.get('reason', '')).border = thin_border
                ws.cell(row=row, column=10, value=status_labels.get(req['status'], '')).border = thin_border
                ws.cell(row=row, column=11, value=req.get('requested_at', '')[:10] if req.get('requested_at') else '').border = thin_border

        elif export_type == "compliance":
            ws.title = f"年次有給休暇管理簿_{year}"
            headers = ["社員番号", "氏名", "基準日", "付与日数", "取得日数", "残日数", "5日義務達成"]

            from agents.compliance import get_compliance
            compliance = get_compliance()
            entries = compliance.generate_annual_ledger(year)
            five_day_results = compliance.check_all_5_day_compliance(year)

            # 5日義務チェック結果をマップ
            compliance_map = {}
            for check in five_day_results.get('checks', []):
                compliance_map[check.employee_num] = check.status.value

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row, entry in enumerate(entries, 2):
                status = compliance_map.get(entry.employee_num, 'unknown')
                status_label = {'compliant': '達成', 'at_risk': '要注意', 'non_compliant': '未達成'}.get(status, '-')

                ws.cell(row=row, column=1, value=entry.employee_num).border = thin_border
                ws.cell(row=row, column=2, value=entry.employee_name).border = thin_border
                ws.cell(row=row, column=3, value=entry.grant_date).border = thin_border
                ws.cell(row=row, column=4, value=entry.granted_days).border = thin_border
                ws.cell(row=row, column=5, value=entry.used_days).border = thin_border
                ws.cell(row=row, column=6, value=entry.remaining_days).border = thin_border
                cell = ws.cell(row=row, column=7, value=status_label)
                cell.border = thin_border
                if status == 'non_compliant':
                    cell.fill = PatternFill(start_color="f87171", end_color="f87171", fill_type="solid")
                elif status == 'at_risk':
                    cell.fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")

        # 列幅自動調整
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)) * 1.5)
            ws.column_dimensions[column_letter].width = max(12, min(50, max_length))

        # ファイル保存
        filename = f"{export_type}_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = UPLOAD_DIR / filename
        wb.save(filepath)

        return {
            "status": "success",
            "message": f"{export_type}データをエクスポートしました",
            "filename": filename,
            "path": str(filepath)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === ANALYTICS ENDPOINTS ===

@app.get("/api/analytics/dashboard/{year}")
async def get_dashboard_analytics(year: int):
    """
    ダッシュボード用の詳細分析データを取得。
    """
    try:
        from collections import defaultdict

        # 基本データ
        employees = database.get_employees(year=year)
        genzai = database.get_genzai()
        ukeoi = database.get_ukeoi()
        requests = database.get_leave_requests(year=year)

        # 部門別集計
        dept_stats = defaultdict(lambda: {'count': 0, 'total_used': 0, 'total_granted': 0})
        for emp in employees:
            dept = emp.get('haken') or '未分類'
            dept_stats[dept]['count'] += 1
            dept_stats[dept]['total_used'] += emp.get('used', 0)
            dept_stats[dept]['total_granted'] += emp.get('granted', 0)

        # 消化率分布
        rate_distribution = {'0-25': 0, '26-50': 0, '51-75': 0, '76-100': 0}
        for emp in employees:
            rate = emp.get('usage_rate', 0)
            if rate <= 25:
                rate_distribution['0-25'] += 1
            elif rate <= 50:
                rate_distribution['26-50'] += 1
            elif rate <= 75:
                rate_distribution['51-75'] += 1
            else:
                rate_distribution['76-100'] += 1

        # 月別トレンド
        monthly = database.get_monthly_usage_summary(year)
        monthly_trend = []
        for month in range(1, 13):
            monthly_trend.append({
                'month': month,
                'total_days': monthly.get(month, {}).get('total_days', 0),
                'employee_count': monthly.get(month, {}).get('employee_count', 0)
            })

        # 申請統計
        request_stats = {
            'total': len(requests),
            'pending': len([r for r in requests if r['status'] == 'PENDING']),
            'approved': len([r for r in requests if r['status'] == 'APPROVED']),
            'rejected': len([r for r in requests if r['status'] == 'REJECTED']),
            'by_type': defaultdict(int)
        }
        for req in requests:
            request_stats['by_type'][req.get('leave_type', 'full')] += 1

        # 従業員タイプ別
        genzai_nums = {e['employee_num'] for e in genzai}
        ukeoi_nums = {e['employee_num'] for e in ukeoi}

        type_stats = {
            '派遣': {'count': 0, 'used': 0},
            '請負': {'count': 0, 'used': 0},
            'その他': {'count': 0, 'used': 0}
        }
        for emp in employees:
            emp_num = emp.get('employee_num', '')
            if emp_num in genzai_nums:
                type_stats['派遣']['count'] += 1
                type_stats['派遣']['used'] += emp.get('used', 0)
            elif emp_num in ukeoi_nums:
                type_stats['請負']['count'] += 1
                type_stats['請負']['used'] += emp.get('used', 0)
            else:
                type_stats['その他']['count'] += 1
                type_stats['その他']['used'] += emp.get('used', 0)

        # Top 10使用者
        top_users = sorted(employees, key=lambda x: x.get('used', 0), reverse=True)[:10]

        # 残日数の多い順
        high_balance = sorted(employees, key=lambda x: x.get('balance', 0), reverse=True)[:10]

        return {
            "status": "success",
            "year": year,
            "summary": {
                "total_employees": len(employees),
                "total_granted": sum(e.get('granted', 0) for e in employees),
                "total_used": sum(e.get('used', 0) for e in employees),
                "total_balance": sum(e.get('balance', 0) for e in employees),
                "average_rate": round(sum(e.get('usage_rate', 0) for e in employees) / len(employees), 1) if employees else 0
            },
            "department_stats": [{'name': k, **v} for k, v in sorted(dept_stats.items(), key=lambda x: x[1]['total_used'], reverse=True)],
            "rate_distribution": rate_distribution,
            "monthly_trend": monthly_trend,
            "request_stats": {**request_stats, 'by_type': dict(request_stats['by_type'])},
            "type_stats": type_stats,
            "top_users": [{'name': e['name'], 'employee_num': e['employee_num'], 'used': e['used']} for e in top_users],
            "high_balance": [{'name': e['name'], 'employee_num': e['employee_num'], 'balance': e['balance']} for e in high_balance]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# === MONTHLY REPORTS ENDPOINTS ===

@app.get("/api/reports/custom")
async def get_custom_report(start_date: str, end_date: str):
    """
    カスタム期間レポート - 任意の日付範囲
    例: /api/reports/custom?start_date=2025-01-16&end_date=2025-02-20
    """
    try:
        import calendar
        from collections import defaultdict

        # Parse dates
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        if end < start:
            raise HTTPException(status_code=400, detail="終了日は開始日より後である必要があります")

        # Get all usage details in date range
        all_usage = []

        # Get unique year-month combinations in range
        current = start
        while current <= end:
            usage = database.get_yukyu_usage_details(year=current.year, month=current.month)
            for u in usage:
                use_date = u.get('use_date', '')
                if use_date and start_date <= use_date <= end_date:
                    all_usage.append(u)
            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        # Get approved requests in range
        approved_requests = database.get_leave_requests(status='APPROVED')
        for req in approved_requests:
            req_start = req.get('start_date', '')
            req_end = req.get('end_date', '')
            if req_start and req_end:
                if req_start <= end_date and req_end >= start_date:
                    exists = any(
                        u.get('employee_num') == req.get('employee_num') and
                        u.get('use_date') == req_start
                        for u in all_usage
                    )
                    if not exists:
                        all_usage.append({
                            'employee_num': req.get('employee_num'),
                            'name': req.get('employee_name'),
                            'use_date': req_start,
                            'days_used': req.get('days_requested', 0),
                            'hours_used': req.get('hours_requested', 0),
                            'leave_type': req.get('leave_type', 'full'),
                            'source': 'request'
                        })

        # Employee summary
        employee_summary = defaultdict(lambda: {
            'name': '',
            'total_days': 0,
            'total_hours': 0,
            'dates': [],
            'factory': ''
        })

        for u in all_usage:
            emp_num = u.get('employee_num', '')
            if emp_num:
                employee_summary[emp_num]['name'] = u.get('name', '')
                employee_summary[emp_num]['total_days'] += u.get('days_used', 0) or 1
                employee_summary[emp_num]['total_hours'] += u.get('hours_used', 0) or 0
                employee_summary[emp_num]['dates'].append({
                    'date': u.get('use_date'),
                    'days': u.get('days_used', 0) or 1,
                    'hours': u.get('hours_used', 0),
                    'type': u.get('leave_type', 'full')
                })

        # Add factory info
        genzai = database.get_genzai()
        ukeoi = database.get_ukeoi()
        genzai_map = {e['employee_num']: e.get('dispatch_name', '') for e in genzai}
        ukeoi_map = {e['employee_num']: e.get('contract_business', '') for e in ukeoi}

        for emp_num, data in employee_summary.items():
            data['factory'] = genzai_map.get(emp_num) or ukeoi_map.get(emp_num) or ''

        # Factory summary
        factory_summary = defaultdict(lambda: {'count': 0, 'total_days': 0, 'employees': []})
        for emp_num, data in employee_summary.items():
            factory = data['factory'] or '未分類'
            factory_summary[factory]['count'] += 1
            factory_summary[factory]['total_days'] += data['total_days']
            factory_summary[factory]['employees'].append({
                'employee_num': emp_num,
                'name': data['name'],
                'days': data['total_days'],
                'hours': data['total_hours']
            })

        # Daily summary
        daily_summary = defaultdict(lambda: {'count': 0, 'employees': []})
        for u in all_usage:
            date = u.get('use_date', '')
            if date:
                daily_summary[date]['count'] += 1
                daily_summary[date]['employees'].append(u.get('name', ''))

        # Calculate days in period
        days_in_period = (end - start).days + 1

        return {
            "status": "success",
            "report_period": {
                "start_date": start_date,
                "end_date": end_date,
                "days_in_period": days_in_period,
                "label": f"{start_date} 〜 {end_date} ({days_in_period}日間)"
            },
            "summary": {
                "total_employees": len(employee_summary),
                "total_days": sum(e['total_days'] for e in employee_summary.values()),
                "total_hours": sum(e['total_hours'] for e in employee_summary.values())
            },
            "employees": [
                {
                    "employee_num": emp_num,
                    "name": data['name'],
                    "factory": data['factory'],
                    "total_days": data['total_days'],
                    "total_hours": data['total_hours'],
                    "dates": sorted(data['dates'], key=lambda x: x['date'])
                }
                for emp_num, data in sorted(employee_summary.items(), key=lambda x: x[1]['total_days'], reverse=True)
            ],
            "by_factory": [
                {
                    "factory": factory,
                    "employee_count": data['count'],
                    "total_days": data['total_days'],
                    "employees": data['employees']
                }
                for factory, data in sorted(factory_summary.items(), key=lambda x: x[1]['total_days'], reverse=True)
            ],
            "by_date": [
                {
                    "date": date,
                    "count": data['count'],
                    "employees": data['employees']
                }
                for date, data in sorted(daily_summary.items())
            ]
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=f"日付フォーマットエラー: {str(ve)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/monthly/{year}/{month}")
async def get_monthly_report(year: int, month: int):
    """
    月次レポート (21日〜20日期間)
    例: 2025年1月レポート = 2024年12月21日 〜 2025年1月20日
    """
    try:
        import calendar
        from collections import defaultdict

        # 期間計算: 前月21日 〜 当月20日
        if month == 1:
            start_year = year - 1
            start_month = 12
        else:
            start_year = year
            start_month = month - 1

        start_date = f"{start_year}-{start_month:02d}-21"
        end_date = f"{year}-{month:02d}-20"

        # 使用日詳細を取得 (期間内)
        all_usage = []

        # 前月21日〜末日のデータ
        _, days_in_prev_month = calendar.monthrange(start_year, start_month)
        usage_prev = database.get_yukyu_usage_details(year=start_year, month=start_month)
        for u in usage_prev:
            use_date = u.get('use_date', '')
            if use_date:
                day = int(use_date.split('-')[2])
                if day >= 21:
                    all_usage.append(u)

        # 当月1日〜20日のデータ
        usage_current = database.get_yukyu_usage_details(year=year, month=month)
        for u in usage_current:
            use_date = u.get('use_date', '')
            if use_date:
                day = int(use_date.split('-')[2])
                if day <= 20:
                    all_usage.append(u)

        # 承認済み申請も取得 (期間内)
        approved_requests = database.get_leave_requests(status='APPROVED')
        for req in approved_requests:
            req_start = req.get('start_date', '')
            req_end = req.get('end_date', '')
            if req_start and req_end:
                # 期間が重なるかチェック
                if req_start <= end_date and req_end >= start_date:
                    # 既存の使用日詳細と重複しないか確認
                    exists = any(
                        u.get('employee_num') == req.get('employee_num') and
                        u.get('use_date') == req_start
                        for u in all_usage
                    )
                    if not exists:
                        all_usage.append({
                            'employee_num': req.get('employee_num'),
                            'name': req.get('employee_name'),
                            'use_date': req_start,
                            'days_used': req.get('days_requested', 0),
                            'hours_used': req.get('hours_requested', 0),
                            'leave_type': req.get('leave_type', 'full'),
                            'source': 'request'
                        })

        # 従業員別集計
        employee_summary = defaultdict(lambda: {
            'name': '',
            'total_days': 0,
            'total_hours': 0,
            'dates': [],
            'factory': ''
        })

        for u in all_usage:
            emp_num = u.get('employee_num', '')
            if emp_num:
                employee_summary[emp_num]['name'] = u.get('name', '')
                employee_summary[emp_num]['total_days'] += u.get('days_used', 0) or 1
                employee_summary[emp_num]['total_hours'] += u.get('hours_used', 0) or 0
                employee_summary[emp_num]['dates'].append({
                    'date': u.get('use_date'),
                    'days': u.get('days_used', 0) or 1,
                    'hours': u.get('hours_used', 0),
                    'type': u.get('leave_type', 'full')
                })

        # 派遣先/契約先情報を追加
        genzai = database.get_genzai()
        ukeoi = database.get_ukeoi()
        genzai_map = {e['employee_num']: e.get('dispatch_name', '') for e in genzai}
        ukeoi_map = {e['employee_num']: e.get('contract_business', '') for e in ukeoi}

        for emp_num, data in employee_summary.items():
            data['factory'] = genzai_map.get(emp_num) or ukeoi_map.get(emp_num) or ''

        # 工場別集計
        factory_summary = defaultdict(lambda: {'count': 0, 'total_days': 0, 'employees': []})
        for emp_num, data in employee_summary.items():
            factory = data['factory'] or '未分類'
            factory_summary[factory]['count'] += 1
            factory_summary[factory]['total_days'] += data['total_days']
            factory_summary[factory]['employees'].append({
                'employee_num': emp_num,
                'name': data['name'],
                'days': data['total_days'],
                'hours': data['total_hours']
            })

        # 日別集計
        daily_summary = defaultdict(lambda: {'count': 0, 'employees': []})
        for u in all_usage:
            date = u.get('use_date', '')
            if date:
                daily_summary[date]['count'] += 1
                daily_summary[date]['employees'].append(u.get('name', ''))

        return {
            "status": "success",
            "report_period": {
                "year": year,
                "month": month,
                "start_date": start_date,
                "end_date": end_date,
                "label": f"{year}年{month}月度 ({start_date} 〜 {end_date})"
            },
            "summary": {
                "total_employees": len(employee_summary),
                "total_days": sum(e['total_days'] for e in employee_summary.values()),
                "total_hours": sum(e['total_hours'] for e in employee_summary.values())
            },
            "employees": [
                {
                    "employee_num": emp_num,
                    "name": data['name'],
                    "factory": data['factory'],
                    "total_days": data['total_days'],
                    "total_hours": data['total_hours'],
                    "dates": sorted(data['dates'], key=lambda x: x['date'])
                }
                for emp_num, data in sorted(employee_summary.items(), key=lambda x: x[1]['total_days'], reverse=True)
            ],
            "by_factory": [
                {
                    "factory": factory,
                    "employee_count": data['count'],
                    "total_days": data['total_days'],
                    "employees": data['employees']
                }
                for factory, data in sorted(factory_summary.items(), key=lambda x: x[1]['total_days'], reverse=True)
            ],
            "by_date": [
                {
                    "date": date,
                    "count": data['count'],
                    "employees": data['employees']
                }
                for date, data in sorted(daily_summary.items())
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/monthly-list/{year}")
async def get_monthly_report_list(year: int):
    """
    年間の月次レポート一覧 (各月の21日〜20日期間のサマリー)
    """
    try:
        reports = []

        for month in range(1, 13):
            # 期間計算
            if month == 1:
                start_year = year - 1
                start_month = 12
            else:
                start_year = year
                start_month = month - 1

            start_date = f"{start_year}-{start_month:02d}-21"
            end_date = f"{year}-{month:02d}-20"

            # 簡易集計
            total_days = 0
            employee_set = set()

            # 前月21日〜末日
            usage_prev = database.get_yukyu_usage_details(year=start_year, month=start_month)
            for u in usage_prev:
                use_date = u.get('use_date', '')
                if use_date:
                    day = int(use_date.split('-')[2])
                    if day >= 21:
                        total_days += u.get('days_used', 0) or 1
                        employee_set.add(u.get('employee_num'))

            # 当月1日〜20日
            usage_current = database.get_yukyu_usage_details(year=year, month=month)
            for u in usage_current:
                use_date = u.get('use_date', '')
                if use_date:
                    day = int(use_date.split('-')[2])
                    if day <= 20:
                        total_days += u.get('days_used', 0) or 1
                        employee_set.add(u.get('employee_num'))

            reports.append({
                "month": month,
                "label": f"{year}年{month}月度",
                "period": f"{start_date} 〜 {end_date}",
                "employee_count": len(employee_set),
                "total_days": total_days
            })

        return {
            "status": "success",
            "year": year,
            "reports": reports
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/analytics/predictions/{year}")
async def get_predictions(year: int):
    """
    年末までの消化率予測を計算。
    """
    try:
        current_month = datetime.now().month
        remaining_months = 12 - current_month

        employees = database.get_employees(year=year)
        monthly = database.get_monthly_usage_summary(year)

        # 過去月の平均使用日数
        past_months_usage = [monthly.get(m, {}).get('total_days', 0) for m in range(1, current_month + 1)]
        avg_monthly_usage = sum(past_months_usage) / len(past_months_usage) if past_months_usage else 0

        predictions = []
        for emp in employees:
            # 個人の月平均使用率
            emp_monthly_avg = emp.get('used', 0) / current_month if current_month > 0 else 0

            # 年末予測
            predicted_used = emp.get('used', 0) + (emp_monthly_avg * remaining_months)
            predicted_rate = (predicted_used / emp.get('granted', 1)) * 100 if emp.get('granted', 0) > 0 else 0

            # 5日達成予測
            will_meet_5day = predicted_used >= 5

            predictions.append({
                'employee_num': emp.get('employee_num'),
                'name': emp.get('name'),
                'current_used': emp.get('used', 0),
                'predicted_used': round(predicted_used, 1),
                'predicted_rate': round(min(predicted_rate, 100), 1),
                'will_meet_5day': will_meet_5day,
                'days_needed': max(0, 5 - emp.get('used', 0))
            })

        # 5日未達成リスク者
        at_risk = [p for p in predictions if not p['will_meet_5day'] and p['days_needed'] > 0]

        return {
            "status": "success",
            "year": year,
            "current_month": current_month,
            "remaining_months": remaining_months,
            "avg_monthly_usage": round(avg_monthly_usage, 1),
            "predictions": predictions,
            "at_risk_count": len(at_risk),
            "at_risk_employees": sorted(at_risk, key=lambda x: x['days_needed'], reverse=True)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# FISCAL YEAR ENDPOINTS (年度管理)
# ============================================

@app.get("/api/fiscal/config")
async def get_fiscal_configuration():
    """Retorna configuración del año fiscal"""
    current_year = datetime.now().year
    current_month = datetime.now().month
    start_date, end_date = get_fiscal_period(current_year, current_month)

    logger.info("Fiscal config requested")

    return {
        "status": "success",
        "config": FISCAL_CONFIG,
        "grant_table": GRANT_TABLE,
        "current_period": {
            "year": current_year,
            "month": current_month,
            "start_date": start_date,
            "end_date": end_date
        }
    }


@app.post("/api/fiscal/process-carryover")
async def process_carryover(from_year: int, to_year: int):
    """
    Procesa carry-over de fin de año fiscal.
    Copia balances no usados y elimina registros vencidos.
    """
    try:
        if to_year <= from_year:
            raise HTTPException(status_code=400, detail="to_year must be greater than from_year")

        stats = process_year_end_carryover(from_year, to_year)
        logger.info(f"Carryover processed: {from_year} -> {to_year}, stats: {stats}")

        return {
            "status": "success",
            "message": f"Carry-over procesado: {from_year} → {to_year}",
            "stats": stats
        }
    except Exception as e:
        logger.error(f"Carryover error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fiscal/balance-breakdown/{employee_num}")
async def get_balance_breakdown(employee_num: str, year: int = None):
    """
    Obtiene desglose de balance por año de origen (para uso FIFO).
    Muestra qué días son del año actual y cuáles del anterior.
    """
    if not year:
        year = datetime.now().year

    breakdown = get_employee_balance_breakdown(employee_num, year)
    return {"status": "success", "data": breakdown}


@app.get("/api/fiscal/expiring-soon")
async def get_expiring_balances(year: int = None):
    """
    Lista empleados con días próximos a expirar.
    Alerta de días que vencerán al fin del año fiscal.
    """
    if not year:
        year = datetime.now().year

    expiring = check_expiring_soon(year)
    total_expiring = sum(e['expiring_days'] for e in expiring)

    return {
        "status": "success",
        "year": year,
        "employees_count": len(expiring),
        "total_expiring_days": round(total_expiring, 1),
        "data": expiring
    }


@app.get("/api/fiscal/5day-compliance/{year}")
async def get_compliance_report(year: int):
    """
    Verifica cumplimiento de la obligación de 5日取得.
    Empleados con 10+ días deben usar mínimo 5.
    """
    compliance = check_5day_compliance(year)
    logger.info(f"5-day compliance check for {year}: {compliance['compliance_rate']}%")

    return {"status": "success", **compliance}


@app.get("/api/fiscal/grant-recommendation/{employee_num}")
async def get_grant_rec(employee_num: str):
    """
    Calcula días a otorgar basado en antigüedad del empleado.
    Usa la tabla de otorgamiento de la Ley Laboral Japonesa.
    """
    recommendation = get_grant_recommendation(employee_num)

    if 'error' in recommendation:
        raise HTTPException(status_code=404, detail=recommendation['error'])

    return {"status": "success", "data": recommendation}


@app.post("/api/fiscal/apply-fifo-deduction")
async def apply_deduction(employee_num: str, days: float, year: int = None):
    """
    Aplica deducción de días usando lógica LIFO.
    Usa primero los días más nuevos (recientes).
    """
    if not year:
        year = datetime.now().year

    try:
        result = apply_lifo_deduction(employee_num, days, year)
        logger.info(f"LIFO deduction: {employee_num}, {days} days, result: {result}")

        return {"status": "success", **result}
    except Exception as e:
        logger.error(f"LIFO deduction error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# EXCEL EXPORT ENDPOINTS (Excel出力)
# ============================================

@app.post("/api/export/approved-requests")
async def export_approved_requests(year: int, month: int = None):
    """Exporta solicitudes aprobadas a Excel"""
    try:
        filepath = create_approved_requests_excel(year, month)
        filename = os.path.basename(filepath)
        logger.info(f"Exported approved requests: {filename}")

        return {
            "status": "success",
            "filename": filename,
            "download_url": f"/api/export/download/{filename}"
        }
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export/monthly-report")
async def export_monthly_report(year: int, month: int):
    """Exporta reporte mensual (21日〜20日) a Excel"""
    try:
        filepath = create_monthly_report_excel(year, month)
        filename = os.path.basename(filepath)
        logger.info(f"Exported monthly report: {filename}")

        return {
            "status": "success",
            "filename": filename,
            "download_url": f"/api/export/download/{filename}"
        }
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export/annual-ledger")
async def export_annual_ledger(year: int):
    """
    Exporta libro de gestión anual (年次有給休暇管理簿).
    Requerido por ley japonesa.
    """
    try:
        filepath = create_annual_ledger_excel(year)
        filename = os.path.basename(filepath)
        logger.info(f"Exported annual ledger: {filename}")

        return {
            "status": "success",
            "filename": filename,
            "download_url": f"/api/export/download/{filename}"
        }
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/download/{filename}")
async def download_export(filename: str):
    """Descarga archivo exportado"""
    # Sanitize filename to prevent path traversal
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(EXPORT_DIR, safe_filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    return FileResponse(
        filepath,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=safe_filename
    )


@app.get("/api/export/files")
async def list_export_files():
    """Lista archivos exportados disponibles"""
    files = get_export_files()
    return {
        "status": "success",
        "count": len(files),
        "files": files
    }


@app.delete("/api/export/cleanup")
async def cleanup_exports(days_to_keep: int = 30, user: dict = Depends(require_admin)):
    """Elimina exportaciones antiguas - requires admin authentication."""
    result = cleanup_old_exports(days_to_keep)
    logger.info(f"Export cleanup by {user['username']}: {result}")
    return {"status": "success", **result}


@app.post("/api/sync/update-master-excel")
async def sync_to_master_excel(year: int):
    """
    Actualiza el archivo Excel maestro con datos de la BD.
    Sincronización bidireccional: BD → Excel
    """
    result = update_master_excel(DEFAULT_EXCEL_PATH, year)

    if result.get("status") == "error":
        logger.error(f"Master Excel update error: {result.get('message')}")
        raise HTTPException(status_code=500, detail=result.get("message"))

    logger.info(f"Master Excel updated: {result}")
    return result


# ============================================
# HEALTH & INFO ENDPOINTS
# ============================================

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "2.0.0"
    }


@app.get("/api/db-status")
async def get_db_status():
    """
    Retorna el estado actual de la base de datos.
    Útil para debugging y verificar que los datos persisten.
    """
    try:
        employees = database.get_employees()
        genzai = database.get_genzai()
        ukeoi = database.get_ukeoi()
        years = database.get_available_years()

        # Check if Excel files exist
        vacation_excel_exists = DEFAULT_EXCEL_PATH.exists()
        registry_excel_exists = EMPLOYEE_REGISTRY_PATH.exists()

        return {
            "status": "success",
            "database": {
                "employees_count": len(employees),
                "genzai_count": len(genzai),
                "ukeoi_count": len(ukeoi),
                "available_years": years,
                "is_empty": len(employees) == 0
            },
            "excel_files": {
                "vacation_excel": {
                    "path": str(DEFAULT_EXCEL_PATH),
                    "exists": vacation_excel_exists
                },
                "employee_registry": {
                    "path": str(EMPLOYEE_REGISTRY_PATH),
                    "exists": registry_excel_exists
                }
            },
            "message": "データは正常に保存されています" if len(employees) > 0 else "データベースは空です - Syncボタンを押してください"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }


@app.get("/api/info")
async def app_info():
    """Application information"""
    return {
        "name": "YuKyuDATA-app",
        "version": "2.0.0",
        "description": "Employee Paid Leave Management System (有給休暇管理システム)",
        "features": [
            "Vacation tracking and balance management",
            "Leave request workflow",
            "Monthly reports (21日〜20日 period)",
            "5-day compliance monitoring",
            "FIFO deduction logic",
            "Year-end carry-over processing",
            "Excel bidirectional sync",
            "Annual ledger generation"
        ],
        "fiscal_config": FISCAL_CONFIG
    }


if __name__ == "__main__":
    logger.info("Starting YuKyuDATA-app server...")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

