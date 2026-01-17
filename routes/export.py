"""
Export Routes
Endpoints de exportacion de datos a Excel
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import FileResponse
from typing import Optional
from datetime import datetime
from pathlib import Path

from .dependencies import (
    get_current_user,
    get_admin_user,
    CurrentUser,
    database,
    logger,
)
from excel_export import (
    create_approved_requests_excel,
    create_monthly_report_excel,
    create_annual_ledger_excel,
    get_export_files,
    cleanup_old_exports,
    EXPORT_DIR
)

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.post("/excel")
async def export_to_excel(
    export_type: str = "employees",
    year: int = None,
    user: CurrentUser = Depends(get_current_user)
):
    """
    Export data to Excel format.
    データをExcel形式でエクスポート。

    export_type: employees, requests, compliance, calendar
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not year:
            year = datetime.now().year

        wb = Workbook()
        ws = wb.active

        # Header styles
        header_fill = PatternFill(start_color="38bdf8", end_color="38bdf8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if export_type == "employees":
            ws.title = f"有給休暇一覧_{year}"
            headers = ["社員番号", "氏名", "派遣先", "付与日数", "使用日数", "残日数", "消化率", "年度"]

            data = database.get_employees(year=year)

            # Headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data
            for row, emp in enumerate(data, 2):
                ws.cell(row=row, column=1, value=emp['employee_num']).border = thin_border
                ws.cell(row=row, column=2, value=emp['name']).border = thin_border
                ws.cell(row=row, column=3, value=emp.get('haken', '')).border = thin_border
                ws.cell(row=row, column=4, value=emp['granted']).border = thin_border
                ws.cell(row=row, column=5, value=emp['used']).border = thin_border
                ws.cell(row=row, column=6, value=emp['balance']).border = thin_border
                ws.cell(row=row, column=7, value=f"{emp.get('usage_rate', 0):.1f}%").border = thin_border
                ws.cell(row=row, column=8, value=emp['year']).border = thin_border

        elif export_type == "requests":
            ws.title = f"休暇申請一覧_{year}"
            headers = ["申請ID", "社員番号", "氏名", "開始日", "終了日", "申請日数", "種別", "状態", "承認者"]

            data = database.get_leave_requests(year=year)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row, req in enumerate(data, 2):
                ws.cell(row=row, column=1, value=req['id']).border = thin_border
                ws.cell(row=row, column=2, value=req['employee_num']).border = thin_border
                ws.cell(row=row, column=3, value=req['employee_name']).border = thin_border
                ws.cell(row=row, column=4, value=req['start_date']).border = thin_border
                ws.cell(row=row, column=5, value=req['end_date']).border = thin_border
                ws.cell(row=row, column=6, value=req['days_requested']).border = thin_border
                ws.cell(row=row, column=7, value=req.get('leave_type', 'full')).border = thin_border
                ws.cell(row=row, column=8, value=req['status']).border = thin_border
                ws.cell(row=row, column=9, value=req.get('approver', '')).border = thin_border

        else:
            raise HTTPException(status_code=400, detail="Unknown export type")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save file
        filename = f"{export_type}_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = EXPORT_DIR / filename
        EXPORT_DIR.mkdir(exist_ok=True)
        wb.save(filepath)

        logger.info(f"Excel export by {user.username}: {filename}")

        return FileResponse(
            path=str(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/approved-requests")
async def export_approved_requests(
    year: int = None,
    month: Optional[int] = None,
    user: CurrentUser = Depends(get_current_user)
):
    """
    Export approved leave requests to Excel.
    承認済み休暇申請をExcelにエクスポート。
    """
    try:
        if not year:
            year = datetime.now().year

        filepath = create_approved_requests_excel(year, month)
        filename = Path(filepath).name

        logger.info(f"Approved requests export by {user.username}: {filename}")

        return FileResponse(
            path=filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except Exception as e:
        logger.error(f"Export approved requests error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/monthly-report")
async def export_monthly_report(
    year: int,
    month: int,
    user: CurrentUser = Depends(get_current_user)
):
    """
    Export monthly report to Excel.
    月次レポートをExcelにエクスポート。
    """
    try:
        filepath = create_monthly_report_excel(year, month)
        filename = Path(filepath).name

        logger.info(f"Monthly report export by {user.username}: {filename}")

        return FileResponse(
            path=filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except Exception as e:
        logger.error(f"Export monthly report error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/annual-ledger")
async def export_annual_ledger(
    year: int,
    user: CurrentUser = Depends(get_current_user)
):
    """
    Export annual ledger (年次有給休暇管理簿) to Excel.
    年次有給休暇管理簿をExcelにエクスポート。
    """
    try:
        filepath = create_annual_ledger_excel(year)
        filename = Path(filepath).name

        logger.info(f"Annual ledger export by {user.username}: {filename}")

        return FileResponse(
            path=filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except Exception as e:
        logger.error(f"Export annual ledger error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/download/{filename}")
async def download_export_file(
    filename: str,
    user: CurrentUser = Depends(get_current_user)
):
    """
    Download an exported file.
    エクスポートファイルをダウンロード。
    """
    try:
        safe_filename = Path(filename).name
        filepath = EXPORT_DIR / safe_filename

        if not filepath.exists():
            raise HTTPException(status_code=404, detail="File not found")

        if not filepath.suffix.lower() in ['.xlsx', '.csv', '.json']:
            raise HTTPException(status_code=400, detail="Invalid file type")

        media_type = {
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.csv': 'text/csv',
            '.json': 'application/json'
        }.get(filepath.suffix.lower(), 'application/octet-stream')

        return FileResponse(
            path=str(filepath),
            media_type=media_type,
            filename=safe_filename
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download export file: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/files")
async def list_export_files(user: CurrentUser = Depends(get_current_user)):
    """
    List available export files.
    エクスポートファイル一覧を取得。
    """
    try:
        files = get_export_files()
        return {
            "status": "success",
            "count": len(files),
            "files": files
        }
    except Exception as e:
        logger.error(f"Failed to list export files: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@router.delete("/cleanup")
async def cleanup_export_files(
    days_to_keep: int = 7,
    user: CurrentUser = Depends(get_admin_user)
):
    """
    Delete old export files.
    古いエクスポートファイルを削除。
    """
    try:
        deleted = cleanup_old_exports(days_to_keep)
        logger.info(f"Export cleanup by {user.username}: {deleted} files deleted")
        return {
            "status": "success",
            "deleted_count": deleted,
            "days_threshold": days_to_keep
        }
    except Exception as e:
        logger.error(f"Failed to cleanup export files: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
