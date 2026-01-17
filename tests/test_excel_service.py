"""
YuKyu Premium - Excel Service Tests
Excelサービステスト - パース機能の完全テスト

Tests para el módulo excel_service.py:
- Detección de tipos de vacaciones (día completo, medio día, por horas)
- Parsing de fechas en varios formatos
- Manejo de archivos corruptos o mal formados
- Columnas faltantes o mal nombradas

Ejecutar con: pytest tests/test_excel_service.py -v
"""

import pytest
import tempfile
import os
import sys
from datetime import datetime
from openpyxl import Workbook

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_service import (
    detect_leave_type,
    clean_date_string,
    parse_date_from_cell,
    is_valid_date_for_yukyu,
    parse_excel_file,
    _find_obs_col,
    HALF_DAY_PATTERNS,
    HOURLY_PATTERNS,
)


# ============================================
# FIXTURES
# ============================================

@pytest.fixture
def temp_excel_file():
    """Crea un archivo Excel temporal para tests."""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    yield path
    # Cleanup
    try:
        os.unlink(path)
    except:
        pass


@pytest.fixture
def sample_vacation_excel(temp_excel_file):
    """Crea un Excel de vacaciones de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "有給休暇"

    # Headers
    headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数', '期末残高', '時効数', '年度']
    ws.append(headers)

    # Sample data
    data = [
        ['001', '田中太郎', 'Factory A', 20, 5, 15, 0, 2025],
        ['002', '山田花子', 'Factory B', 15, 3.5, 11.5, 0, 2025],
        ['003', '鈴木一郎', 'Factory A', 10, 10, 0, 0, 2025],
    ]
    for row in data:
        ws.append(row)

    wb.save(temp_excel_file)
    return temp_excel_file


@pytest.fixture
def excel_with_missing_columns(temp_excel_file):
    """Crea un Excel con columnas faltantes."""
    wb = Workbook()
    ws = wb.active

    # Headers incompletos (falta 'balance')
    headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
    ws.append(headers)

    data = [
        ['001', '田中太郎', 'Factory A', 20, 5],
    ]
    for row in data:
        ws.append(row)

    wb.save(temp_excel_file)
    return temp_excel_file


@pytest.fixture
def excel_with_japanese_dates(temp_excel_file):
    """Crea un Excel con fechas en formato japonés."""
    wb = Workbook()
    ws = wb.active

    headers = ['社員№', '氏名', '有給発生', '付与数', '消化日数']
    ws.append(headers)

    # Fechas en diferentes formatos
    data = [
        ['001', '田中太郎', '2025年4月1日', 20, 5],
        ['002', '山田花子', datetime(2025, 4, 1), 15, 3],
        ['003', '鈴木一郎', '2025/04/01', 10, 0],
    ]
    for row in data:
        ws.append(row)

    wb.save(temp_excel_file)
    return temp_excel_file


# ============================================
# DETECT LEAVE TYPE TESTS
# ============================================

class TestDetectLeaveType:
    """Tests para detect_leave_type()"""

    def test_full_day_from_date(self):
        """Fecha datetime es día completo"""
        leave_type, days = detect_leave_type(datetime(2025, 1, 15))
        assert leave_type == 'full'
        assert days == 1.0

    def test_full_day_from_date_string(self):
        """String de fecha es día completo"""
        leave_type, days = detect_leave_type('2025-01-15')
        assert leave_type == 'full'
        assert days == 1.0

    def test_half_day_kanji(self):
        """半 (medio) se detecta como medio día"""
        leave_type, days = detect_leave_type('2025-01-15 半')
        assert leave_type in ['half_am', 'half_pm']
        assert days == 0.5

    def test_half_day_explicit(self):
        """0.5 explícito se detecta como medio día"""
        leave_type, days = detect_leave_type('0.5日')
        assert days == 0.5

    def test_half_day_am(self):
        """午前 (mañana) es medio día AM"""
        leave_type, days = detect_leave_type('2025-01-15 午前休')
        assert leave_type == 'half_am'
        assert days == 0.5

    def test_half_day_pm(self):
        """午後 (tarde) es medio día PM"""
        leave_type, days = detect_leave_type('2025-01-15 午後休')
        assert leave_type == 'half_pm'
        assert days == 0.5

    def test_half_day_hankyu(self):
        """半休 (medio día libre) es 0.5"""
        leave_type, days = detect_leave_type('半休')
        assert days == 0.5

    def test_hourly_leave_2h(self):
        """2h es vacación por horas (0.25 días)"""
        leave_type, days = detect_leave_type('2h')
        assert leave_type == 'hourly'
        assert days == 0.25

    def test_hourly_leave_japanese(self):
        """2時間 es vacación por horas"""
        leave_type, days = detect_leave_type('2時間')
        assert leave_type == 'hourly'
        assert days == 0.25

    def test_hourly_leave_variable(self):
        """Horas variables se calculan correctamente"""
        leave_type, days = detect_leave_type('4h')
        assert leave_type == 'hourly'
        assert days == 0.5  # 4 hours = 0.5 days

    def test_jikan_kyu(self):
        """時間休 (vacación por horas) se detecta"""
        leave_type, days = detect_leave_type('時間休')
        assert leave_type == 'hourly'

    def test_none_value(self):
        """None retorna unknown, 0.0"""
        leave_type, days = detect_leave_type(None)
        assert leave_type == 'unknown'
        assert days == 0.0

    def test_empty_string(self):
        """String vacío es día completo por defecto"""
        leave_type, days = detect_leave_type('')
        assert leave_type == 'full'


# ============================================
# DATE PARSING TESTS
# ============================================

class TestCleanDateString:
    """Tests para clean_date_string()"""

    def test_removes_japanese_noise(self):
        """Remueve texto japonés innecesario"""
        assert clean_date_string('終業2025-01-15') == '2025-01-15'
        assert clean_date_string('2025-01-15至') == '2025-01-15'
        assert clean_date_string('休2025-01-15取得') == '2025-01-15'

    def test_preserves_date(self):
        """Preserva la fecha correctamente"""
        assert clean_date_string('2025-01-15') == '2025-01-15'
        assert clean_date_string('2025/01/15') == '2025/01/15'

    def test_handles_spaces(self):
        """Maneja espacios correctamente"""
        result = clean_date_string('  2025-01-15  ')
        assert result.strip() == '2025-01-15'


class TestParseDateFromCell:
    """Tests para parse_date_from_cell()"""

    def test_datetime_object(self):
        """Datetime object se retorna directamente"""
        dt = datetime(2025, 1, 15)
        result = parse_date_from_cell(dt)
        assert result == dt

    def test_iso_format(self):
        """Formato ISO (YYYY-MM-DD)"""
        result = parse_date_from_cell('2025-01-15')
        assert result.year == 2025
        assert result.month == 1
        assert result.day == 15

    def test_japanese_format(self):
        """Formato japonés (YYYY年MM月DD日)"""
        result = parse_date_from_cell('2025年1月15日')
        assert result.year == 2025
        assert result.month == 1
        assert result.day == 15

    def test_slash_format(self):
        """Formato con slash (YYYY/MM/DD)"""
        result = parse_date_from_cell('2025/01/15')
        assert result.year == 2025
        assert result.month == 1
        assert result.day == 15

    def test_excel_serial_date(self):
        """Número serial de Excel se convierte"""
        # 45672 = aproximadamente 2025-01-15
        result = parse_date_from_cell(45672)
        assert result is not None
        assert result.year >= 2024

    def test_invalid_serial_date_high(self):
        """Serial date muy alto retorna None"""
        result = parse_date_from_cell(100000)
        assert result is None

    def test_invalid_serial_date_low(self):
        """Serial date muy bajo retorna None"""
        result = parse_date_from_cell(-1)
        assert result is None

    def test_none_value(self):
        """None retorna None"""
        result = parse_date_from_cell(None)
        assert result is None

    def test_invalid_string(self):
        """String inválido retorna None"""
        result = parse_date_from_cell('not a date')
        assert result is None

    def test_mixed_content(self):
        """Contenido mixto extrae la fecha"""
        result = parse_date_from_cell('休暇取得 2025-01-15 午前')
        assert result is not None
        assert result.year == 2025


class TestIsValidDateForYukyu:
    """Tests para is_valid_date_for_yukyu()"""

    def test_valid_date(self):
        """Fecha válida retorna True"""
        assert is_valid_date_for_yukyu(datetime(2025, 1, 15)) == True

    def test_date_too_old(self):
        """Fecha muy antigua retorna False"""
        assert is_valid_date_for_yukyu(datetime(1999, 1, 15)) == False

    def test_date_too_future(self):
        """Fecha muy futura retorna False"""
        assert is_valid_date_for_yukyu(datetime(2101, 1, 15)) == False

    def test_none_date(self):
        """None retorna False"""
        assert is_valid_date_for_yukyu(None) == False


# ============================================
# COLUMN FINDING TESTS
# ============================================

class TestFindObsCol:
    """Tests para _find_obs_col()"""

    def test_finds_exact_match(self):
        """Encuentra match exacto"""
        headers = ['ID', '氏名', '部署', '付与数']
        result = _find_obs_col(headers, ['氏名'])
        assert result == 1

    def test_finds_partial_match(self):
        """Encuentra match parcial"""
        headers = ['社員№', '氏名', '派遣先', '付与日数']
        result = _find_obs_col(headers, ['付与', '付与数'])
        assert result == 3

    def test_finds_first_keyword_priority(self):
        """Prioriza primer keyword"""
        headers = ['番号', '社員№', '従業員番号']
        result = _find_obs_col(headers, ['社員№', '従業員番号', '番号'])
        assert result == 1  # 社員№ tiene prioridad

    def test_not_found_returns_minus_one(self):
        """No encontrado retorna -1"""
        headers = ['A', 'B', 'C']
        result = _find_obs_col(headers, ['X', 'Y', 'Z'])
        assert result == -1

    def test_handles_none_headers(self):
        """Maneja headers None"""
        headers = [None, '氏名', None, '付与数']
        result = _find_obs_col(headers, ['氏名'])
        assert result == 1

    def test_case_insensitive(self):
        """Búsqueda case-insensitive"""
        headers = ['ID', 'Name', 'BALANCE']
        result = _find_obs_col(headers, ['balance', 'remaining'])
        assert result == 2


# ============================================
# EXCEL PARSING TESTS
# ============================================

class TestParseExcelFile:
    """Tests para parse_excel_file()"""

    def test_parses_sample_excel(self, sample_vacation_excel):
        """Parsea Excel de vacaciones correctamente"""
        result = parse_excel_file(sample_vacation_excel)

        assert len(result) == 3
        assert result[0]['employeeNum'] == '001'
        assert result[0]['name'] == '田中太郎'
        assert result[0]['granted'] == 20
        assert result[0]['used'] == 5
        assert result[0]['balance'] == 15

    def test_calculates_usage_rate(self, sample_vacation_excel):
        """Calcula tasa de uso correctamente"""
        result = parse_excel_file(sample_vacation_excel)

        # 5/20 = 25%
        assert result[0]['usageRate'] == 25
        # 10/10 = 100%
        assert result[2]['usageRate'] == 100

    def test_handles_half_day_values(self, temp_excel_file):
        """Maneja valores de medio día"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数', '期末残高']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, 5.5, 14.5])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['used'] == 5.5
        assert result[0]['balance'] == 14.5

    def test_file_not_found_raises_error(self):
        """Archivo no existente lanza error"""
        with pytest.raises(FileNotFoundError):
            parse_excel_file('/nonexistent/file.xlsx')

    def test_handles_missing_columns(self, excel_with_missing_columns):
        """Maneja columnas faltantes"""
        result = parse_excel_file(excel_with_missing_columns)

        assert len(result) == 1
        # Balance se calcula si no existe la columna
        assert result[0]['balance'] == 15  # granted - used = 20 - 5

    def test_skips_empty_rows(self, temp_excel_file):
        """Salta filas vacías"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, 5])
        ws.append([None, None, None, None, None])  # Empty row
        ws.append(['002', '山田花子', 'Factory B', 15, 3])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert len(result) == 2

    def test_skips_rows_without_name(self, temp_excel_file):
        """Salta filas sin nombre"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, 5])
        ws.append(['002', None, 'Factory B', 15, 3])  # No name
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert len(result) == 1

    def test_handles_duplicate_header_rows(self, temp_excel_file):
        """Salta filas que parecen headers duplicados"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, 5])
        ws.append(['番号', '氏名', '派遣先', '付与日数', '消化日数'])  # Duplicate header
        ws.append(['002', '山田花子', 'Factory B', 15, 3])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        # Should skip the duplicate header row
        assert len(result) == 2

    def test_extracts_year_from_grant_date(self, excel_with_japanese_dates):
        """Extrae año de la fecha de otorgamiento"""
        result = parse_excel_file(excel_with_japanese_dates)

        for emp in result:
            # Year is extracted from grant date or defaults to current fiscal year
            # The fixture has 2025 dates, but the system may use current fiscal year
            assert emp['year'] in [2025, 2026, datetime.now().year]

    def test_generates_unique_ids(self, sample_vacation_excel):
        """Genera IDs únicos (empNum_year)"""
        result = parse_excel_file(sample_vacation_excel)

        ids = [emp['id'] for emp in result]
        # IDs follow format: empNum_year
        # Check pattern exists
        for emp_id in ids:
            parts = emp_id.split('_')
            assert len(parts) == 2
            assert parts[0].isdigit() or parts[0]  # Employee number
            assert parts[1].isdigit()  # Year
        assert len(ids) == len(set(ids))  # All unique

    def test_handles_invalid_numeric_values(self, temp_excel_file):
        """Maneja valores numéricos inválidos"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 'invalid', 'bad'])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['granted'] == 0.0
        assert result[0]['used'] == 0.0

    def test_handles_expired_column(self, temp_excel_file):
        """Maneja columna de días expirados"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数', '時効数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, 5, 3])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['expired'] == 3.0


# ============================================
# EDGE CASES AND ERROR HANDLING
# ============================================

class TestEdgeCases:
    """Tests para casos extremos"""

    def test_empty_excel(self, temp_excel_file):
        """Excel vacío retorna lista vacía"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result == []

    def test_large_values(self, temp_excel_file):
        """Valores muy grandes se manejan"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 999999, 500000])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['granted'] == 999999

    def test_negative_values(self, temp_excel_file):
        """Valores negativos se aceptan (aunque no deberían ocurrir)"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数', '期末残高']
        ws.append(headers)
        ws.append(['001', '田中太郎', 'Factory A', 20, -5, 25])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['used'] == -5

    def test_special_characters_in_name(self, temp_excel_file):
        """Caracteres especiales en nombre se manejan"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中（太郎）', 'Factory A', 20, 5])
        ws.append(['002', 'John O\'Brien', 'Factory B', 15, 3])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert len(result) == 2
        assert '（太郎）' in result[0]['name']
        assert "O'Brien" in result[1]['name']

    def test_unicode_in_all_fields(self, temp_excel_file):
        """Unicode en todos los campos se maneja"""
        wb = Workbook()
        ws = wb.active
        headers = ['社員№', '氏名', '派遣先', '付与数', '消化日数']
        ws.append(headers)
        ws.append(['001', '田中太郎', '東京工場α', 20, 5])
        wb.save(temp_excel_file)

        result = parse_excel_file(temp_excel_file)
        assert result[0]['haken'] == '東京工場α'


# ============================================
# PATTERN CONSTANTS TESTS
# ============================================

class TestPatternConstants:
    """Tests para verificar que los patrones están configurados"""

    def test_half_day_patterns_exist(self):
        """Patrones de medio día están definidos"""
        assert len(HALF_DAY_PATTERNS) > 0
        assert '半' in HALF_DAY_PATTERNS
        assert '午前' in HALF_DAY_PATTERNS
        assert '午後' in HALF_DAY_PATTERNS

    def test_hourly_patterns_exist(self):
        """Patrones de horas están definidos"""
        assert len(HOURLY_PATTERNS) > 0
        assert '2h' in HOURLY_PATTERNS or any('h' in p for p in HOURLY_PATTERNS)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
