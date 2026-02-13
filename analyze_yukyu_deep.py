
import openpyxl
from collections import Counter
import sys

FILE_PATH = r"E:\CosasParaAppsJp\有給休暇管理.xlsm"

def analyze_excel():
    print(f"--- ANALYZING: {FILE_PATH} ---")
    
    try:
        wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
        print(f"SHEETS FOUND: {wb.sheetnames}")
        
        target_sheet_name = '作業者データ　有給'
        if target_sheet_name not in wb.sheetnames:
            print(f"WARNING: Sheet '{target_sheet_name}' not found. Using first sheet.")
            target_sheet_name = wb.sheetnames[0]
            
        sheet = wb[target_sheet_name]
        
        # 1. Analyze Headers (Rows 1-7) to find exact row
        print("\n--- FINDING HEADER ROW ---")
        header_row_idx = -1
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_str = [str(c) if c is not None else "" for c in row]
            print(f"Row {i}: {row_str[:8]}...") 
            if any("氏名" in rs or "名前" in rs for rs in row_str):
                header_row_idx = i
                print(f"  -> Header found at Row {header_row_idx}")
                # Print full header row
                print(f"  -> FULL HEADER: {row_str}")
                break
        
        if header_row_idx == -1:
            print("  -> Header NOT found in first 10 rows.")
            header_row_idx = 1

        # 2. Hidden Columns
        print("\n--- HIDDEN COLUMNS ---")
        try:
            for col_idx, col_dim in sheet.column_dimensions.items():
                if col_dim.hidden:
                    print(f"  Column {col_idx} is HIDDEN")
        except:
             print("  Could not access column_dimensions (might be openpyxl limitaion on read-only)")

        # 3. Scan Usage Columns (Assumed R=18 to BE=57 based on previous knowledge)
        # We will scan broader range to be safe: Col 15 (O) to 80 (CB)
        print("\n--- CELL VALUE ANALYSIS (Cols 15-80) ---")
        
        text_values = Counter()
        date_count = 0
        empty_count = 0
        total_cells = 0
        
        start_row = header_row_idx + 1
        print(f"Scanning data from Row {start_row}...")

        # Columns 15 to 80 (1-based index)
        for row in sheet.iter_rows(min_row=start_row, max_row=500, min_col=15, max_col=80, values_only=True):
            for cell in row:
                total_cells += 1
                if cell is None:
                    empty_count += 1
                    continue
                
                val_str = str(cell).strip()
                if val_str == "": 
                    empty_count += 1
                    continue
                
                # Check formatting/content
                is_date = False
                # Simple check for date-like strings
                if '-' in val_str or '/' in val_str or '202' in val_str:
                     is_date = True
                
                if isinstance(cell, (int, float)) and cell > 40000: # Excel date serial
                     is_date = True

                if is_date:
                    date_count += 1
                else:
                    text_values[val_str] += 1

        print(f"\nTotal Cells Scanned: {total_cells}")
        print(f"Empty Cells: {empty_count}")
        print(f"Likely Dates: {date_count}")
        print("\n--- NON-DATE TEXT VALUES FOUND (Top 50) ---")
        for val, count in text_values.most_common(50):
            print(f"'{val}': {count}")

    except Exception as e:
        print(f"CRITICAL ERROR: {e}")

if __name__ == "__main__":
    analyze_excel()
