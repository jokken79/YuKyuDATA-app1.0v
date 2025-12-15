
import openpyxl
import os

file_path = r"D:\YuKyuDATA-app\有給休暇管理.xlsm"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.worksheets[0]

# Logic from excel_service.py
header_row_idx = 1
headers = []
found_headers = False
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    row_str = [str(c) if c is not None else "" for c in row]
    if any("氏名" in c or "名前" in c for c in row_str):
        headers = row
        header_row_idx = i
        found_headers = True
        break

if not found_headers:
    headers = [cell.value for cell in sheet[1]]
    header_row_idx = 1

def _find_obs_col(headers, keywords):
    for i, h in enumerate(headers):
        if h is None: continue
        h_str = str(h).lower()
        for k in keywords:
            if k in h_str:
                return i
    return -1

cols = {
    'empNum': _find_obs_col(headers, ['従業員番号', '社員番号', '番号', 'id', 'no', '№']),
    'name': _find_obs_col(headers, ['氏名', '名前', 'name']),
}

print(f"Header Row: {header_row_idx}")
print(f"Name Col Index: {cols['name']}")

parsed_rows = []
for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
    if all(c is None for c in row):
        continue
    
    name = str(row[cols['name']]) if cols['name'] != -1 and row[cols['name']] is not None else "Unknown"
    
    if name == "Unknown":
        continue

    parsed_rows.append(f"Row {i}: {name}")

with open('d:/YuKyuDATA-app/excel_debug.txt', 'w', encoding='utf-8') as f:
    f.write(f"Total Parsed: {len(parsed_rows)}\n")
    for r in parsed_rows:
        f.write(r + "\n")
