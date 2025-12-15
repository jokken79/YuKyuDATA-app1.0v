import openpyxl

FILE_PATH = r"D:\YuKyuDATA-app\有給休暇管理.xlsm"

try:
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
    print("Sheet Names:", wb.sheetnames)
    print(f"Active Sheet: {wb.active.title}")
    
    first_sheet_name = wb.sheetnames[0]
    sheet = wb[first_sheet_name]
    print(f"First Sheet Name: {first_sheet_name}")
    
    # Check first few rows of the FIRST sheet
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        print(f"Row {i+1}: {row}")
        
except Exception as e:
    print(f"Error: {e}")
